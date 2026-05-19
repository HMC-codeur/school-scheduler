from __future__ import annotations

import argparse
import os
import sys
from contextlib import contextmanager
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from backend.evals.excel.checks import ExcelAnalysisAdapter, apply_case_checks
from backend.evals.excel.readiness import calculate_readiness
from backend.evals.excel.reporting import print_terminal_summary, write_json_report, write_markdown_report
from backend.evals.excel.schemas import ExpectedCase, load_expected_cases
from backend.services.imports.excel_mvp.pipeline import analyze_excel_content


ROOT = Path(__file__).resolve().parents[3]
DEFAULT_EXPECTED_DIR = Path(__file__).resolve().parent / "expected"
DEFAULT_REPORT_DIR = Path(__file__).resolve().parent / "reports"


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser(description="Run Excel Intelligence v2 eval suite.")
    parser.add_argument("--case", dest="case_id")
    parser.add_argument("--fixtures", type=Path)
    parser.add_argument("--expected-dir", type=Path, default=DEFAULT_EXPECTED_DIR)
    parser.add_argument("--report-dir", type=Path, default=DEFAULT_REPORT_DIR)
    parser.add_argument("--engine", default="v2")
    parser.add_argument("--fail-on-warning", action="store_true")
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args(argv)
    if args.fixtures:
        report = run_fixture_suite(
            fixtures_dir=args.fixtures,
            engine=args.engine,
            include_raw=args.debug,
            fail_on_warning=args.fail_on_warning,
        )
        print_fixture_summary(report)
        return fixture_exit_code(report, fail_on_warning=args.fail_on_warning)

    report, json_path, md_path = run_eval_suite(
        expected_dir=args.expected_dir,
        report_dir=args.report_dir,
        engine=args.engine,
        case_id=args.case_id,
        include_raw=args.debug,
        fail_on_warning=args.fail_on_warning,
    )
    print_terminal_summary(report, json_path, md_path)
    return 0 if report["status"] == "passed" else 1


def run_fixture_suite(
    *,
    fixtures_dir: Path,
    engine: str = "v2",
    include_raw: bool = False,
    fail_on_warning: bool = False,
) -> dict[str, Any]:
    files = discover_fixture_files(fixtures_dir)
    cases: list[dict[str, Any]] = []
    with temporary_excel_mode(engine):
        for path in files:
            cases.append(run_fixture_file(path, engine=engine, include_raw=include_raw))
    warnings_total = sum(len(case.get("warnings", [])) for case in cases)
    blocking_total = sum(len(case.get("blocking_diagnostics", [])) for case in cases)
    crashes = [case for case in cases if case.get("crashed")]
    failed_warning_status = fail_on_warning and warnings_total > 0
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "engine": engine,
        "fixtures_dir": str(fixtures_dir),
        "status": "failed" if crashes or failed_warning_status else "passed",
        "summary": {
            "files_tested": len(cases),
            "files_passed": sum(1 for case in cases if case["status"] == "passed"),
            "files_failed": sum(1 for case in cases if case["status"] == "failed"),
            "warnings_total": warnings_total,
            "blocking_errors_total": blocking_total,
            "crashes_total": len(crashes),
        },
        "cases": cases,
    }


def discover_fixture_files(fixtures_dir: Path) -> list[Path]:
    if not fixtures_dir.exists():
        raise FileNotFoundError(f"Fixtures directory not found: {fixtures_dir}")
    if not fixtures_dir.is_dir():
        raise NotADirectoryError(f"Fixtures path is not a directory: {fixtures_dir}")
    return sorted(
        path
        for path in fixtures_dir.rglob("*.xlsx")
        if not path.name.startswith("00_MANIFEST")
    )


def run_fixture_file(path: Path, *, engine: str = "v2", include_raw: bool = False) -> dict[str, Any]:
    try:
        result = analyze_excel_content(path.read_bytes(), filename=path.name)
        normalized = ExcelAnalysisAdapter(result, include_raw=include_raw).normalized()
        blocking = _collect_diagnostics(result, "blocking")
        warnings = _collect_diagnostics(result, "warning")
        human_corrections = _collect_human_corrections(result)
        return {
            "file_name": path.name,
            "file_path": str(path),
            "file_size": path.stat().st_size,
            "status": "failed" if blocking else "passed",
            "crashed": False,
            "engine_used": normalized.get("engine_used"),
            "sheets": normalized.get("sheets", []),
            "sheet_names": [sheet.get("sheet_name") for sheet in normalized.get("sheets", [])],
            "detected_formats": _unique_strings(sheet.get("actual_format") for sheet in normalized.get("sheets", [])),
            "summary": normalized.get("summary", {}),
            "blocking_diagnostics": blocking,
            "warnings": warnings,
            "human_corrections": human_corrections,
        }
    except Exception as exc:
        return {
            "file_name": path.name,
            "file_path": str(path),
            "status": "failed",
            "crashed": True,
            "engine_used": engine,
            "sheets": [],
            "sheet_names": [],
            "detected_formats": [],
            "summary": {},
            "blocking_diagnostics": [],
            "warnings": [],
            "human_corrections": [],
            "error": f"{type(exc).__name__}: {exc}",
        }


def fixture_exit_code(report: dict[str, Any], *, fail_on_warning: bool = False) -> int:
    summary = report.get("summary", {})
    if int(summary.get("crashes_total") or 0) > 0:
        return 1
    if fail_on_warning and int(summary.get("warnings_total") or 0) > 0:
        return 1
    return 0


def print_fixture_summary(report: dict[str, Any]) -> None:
    summary = report["summary"]
    print("Excel Fixtures Eval")
    print(f"Engine: {report['engine']}")
    print(f"Fixtures: {report['fixtures_dir']}")
    print(f"Status: {report['status'].upper()}")
    print("")
    for case in report["cases"]:
        print(f"{case['file_name']}: {case['status'].upper()}")
        print(f"  Sheets: {', '.join(str(item) for item in case.get('sheet_names', [])) or 'none'}")
        print(f"  Detected types: {', '.join(case.get('detected_formats', [])) or 'none'}")
        if case.get("summary"):
            print(f"  Counts: {_format_counts(case['summary'])}")
        if case.get("blocking_diagnostics"):
            print("  Blocking diagnostics:")
            for item in case["blocking_diagnostics"][:8]:
                print(f"  - {_diagnostic_line(item)}")
        if case.get("warnings"):
            print("  Warnings:")
            for item in case["warnings"][:8]:
                print(f"  - {_diagnostic_line(item)}")
        if case.get("human_corrections"):
            print("  Human corrections requested:")
            for item in case["human_corrections"][:5]:
                print(f"  - {item}")
        if case.get("error"):
            print(f"  Exception: {case['error']}")
        print("")
    print("Global summary:")
    print(f"- Files tested: {summary['files_tested']}")
    print(f"- Files passed: {summary['files_passed']}")
    print(f"- Files failed: {summary['files_failed']}")
    print(f"- Total warnings: {summary['warnings_total']}")
    print(f"- Total blocking errors: {summary['blocking_errors_total']}")
    print(f"- Crashes: {summary['crashes_total']}")


def _collect_diagnostics(result: dict[str, Any], severity: str) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    diagnostics = result.get("diagnostics") or {}
    if isinstance(diagnostics, dict):
        key = "warnings" if severity == "warning" else severity
        items.extend(item for item in diagnostics.get(key, []) if isinstance(item, dict))
    items.extend(item for item in result.get("global_diagnostics", []) if isinstance(item, dict) and item.get("severity") == severity)
    for sheet in result.get("sheets", []):
        sheet_name = sheet.get("sheet_name")
        for item in sheet.get("diagnostics", []) or []:
            if isinstance(item, dict) and item.get("severity") == severity:
                copied = dict(item)
                copied.setdefault("sheet_name", sheet_name)
                items.append(copied)
    return items


def _collect_human_corrections(result: dict[str, Any]) -> list[str]:
    corrections: list[str] = []
    if result.get("needs_human_mapping"):
        corrections.append("Mapping humain requis pour au moins une feuille.")
    if result.get("needs_human_validation"):
        corrections.append("Validation humaine requise.")
    for question in result.get("validation_questions", []) or []:
        if isinstance(question, dict):
            text = question.get("question") or question.get("message") or question.get("title")
            if text:
                corrections.append(str(text))
        else:
            corrections.append(str(question))
    for sheet in result.get("sheets", []) or []:
        message = sheet.get("human_message")
        if message:
            corrections.append(f"{sheet.get('sheet_name')}: {message}")
    return _unique_strings(corrections)


def _diagnostic_line(item: dict[str, Any]) -> str:
    code = item.get("code") or item.get("id") or item.get("title") or "diagnostic"
    message = item.get("message") or item.get("detail") or item.get("description") or ""
    sheet = item.get("sheet_name")
    prefix = f"{sheet}: " if sheet else ""
    return f"{prefix}{code} - {message}" if message else f"{prefix}{code}"


def _format_counts(summary: dict[str, Any]) -> str:
    keys = ("requirements_detected", "scheduled_lessons_detected", "teacher_availability_detected", "classes_detected", "teachers_detected", "subjects_detected")
    parts = [f"{key}={summary[key]}" for key in keys if key in summary]
    return ", ".join(parts) if parts else "n/a"


def _unique_strings(values: Any) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value is None:
            continue
        text = str(value)
        if text not in seen:
            seen.add(text)
            result.append(text)
    return result


def run_eval_suite(
    *,
    expected_dir: Path = DEFAULT_EXPECTED_DIR,
    report_dir: Path = DEFAULT_REPORT_DIR,
    engine: str = "v2",
    case_id: str | None = None,
    include_raw: bool = False,
    fail_on_warning: bool = False,
) -> tuple[dict[str, Any], Path, Path]:
    cases = load_expected_cases(expected_dir)
    if case_id:
        cases = [case for case in cases if case.case_id == case_id]
    files = inventory_files(cases)
    case_reports: list[dict[str, Any]] = []
    with temporary_excel_mode(engine):
        for case in cases:
            case_reports.append(run_case(case, engine=engine, include_raw=include_raw))
    gates = apply_global_gates(case_reports, engine=engine)
    readiness = calculate_readiness(case_reports, gates)
    report = build_report(engine, case_reports, gates, readiness, files, fail_on_warning=fail_on_warning)
    json_path = write_json_report(report, report_dir)
    md_path = write_markdown_report(report, report_dir)
    return report, json_path, md_path


def run_case(case: ExpectedCase, *, engine: str = "v2", include_raw: bool = False) -> dict[str, Any]:
    path = case.resolved_file_path(ROOT)
    if case.synthetic and not path.exists():
        create_synthetic_core_workbook(path)
    if not path.exists():
        return {
            "case_id": case.case_id,
            "description": case.description,
            "file_path": case.file_path,
            "synthetic": case.synthetic,
            "status": "skipped" if case.skip_if_file_missing else "failed",
            "skip_reason": "file_missing" if case.skip_if_file_missing else None,
            "error": None if case.skip_if_file_missing else f"File not found: {path}",
            "checks": [],
            "sheets": [],
        }
    try:
        result = analyze_excel_content(path.read_bytes(), filename=path.name)
        normalized = ExcelAnalysisAdapter(result, include_raw=include_raw).normalized()
        checks = apply_case_checks(case, normalized)
        failed_blocking = [item for item in checks if not item["passed"] and item["severity"] == "blocking"]
        return {
            "case_id": case.case_id,
            "description": case.description,
            "file_path": case.file_path,
            "file_size": path.stat().st_size,
            "synthetic": case.synthetic,
            "status": "failed" if failed_blocking else "passed",
            "engine_used": normalized.get("engine_used"),
            "checks": checks,
            "sheets": normalized.get("sheets", []),
            "summary": normalized.get("summary", {}),
        }
    except Exception as exc:
        return {
            "case_id": case.case_id,
            "description": case.description,
            "file_path": case.file_path,
            "synthetic": case.synthetic,
            "status": "failed",
            "error": f"{type(exc).__name__}: {exc}",
            "checks": [],
            "sheets": [],
        }


def apply_global_gates(cases: list[dict[str, Any]], *, engine: str) -> list[dict[str, Any]]:
    executed = [case for case in cases if case.get("status") != "skipped"]
    sheets = [sheet for case in executed for sheet in case.get("sheets", [])]

    def gate(gate_id: str, title: str, passed: bool, message: str, evidence: dict[str, Any] | None = None, severity: str = "blocking") -> dict[str, Any]:
        return {"id": gate_id, "title": title, "severity": severity, "passed": passed, "message": message, "evidence": evidence or {}, "related_cases": [case["case_id"] for case in executed]}

    fake_lessons = [fake for sheet in sheets for fake in sheet.get("fake_availability_lessons", [])]
    fake_requirements = [fake for sheet in sheets for fake in sheet.get("fake_requirements", [])]
    availability_sheets = [sheet for sheet in sheets if sheet.get("actual_format") == "availability_grid" or "availability" in str(sheet.get("sheet_name", "")).lower() or "זמינות" in str(sheet.get("sheet_name", ""))]
    requirement_sheets = [sheet for sheet in sheets if sheet.get("sheet_name") in {"01 Besoins MIX", "requirements", "synthetic_requirements"} or sheet.get("actual_format") == "requirements_table"]
    schedule_sheets = [sheet for sheet in sheets if sheet.get("sheet_name") in {"02 מערכת שעות גריד", "synthetic_schedule"} or sheet.get("actual_format") == "schedule_grid"]
    noise_sheets = [sheet for sheet in sheets if "noise" in str(sheet.get("sheet_name", "")).lower() or "Unknown Noise" in str(sheet.get("sheet_name", ""))]
    constraints = [sheet for sheet in sheets if "contrainte" in str(sheet.get("sheet_name", "")).lower() or "constraint" in str(sheet.get("sheet_name", "")).lower()]
    mixed = [sheet for sheet in sheets if "mixed" in str(sheet.get("sheet_name", "")).lower() or "liste" in str(sheet.get("sheet_name", "")).lower()]
    v1_default_ok = _v1_default_preserved()

    return [
        gate("no_availability_markers_as_lessons", "No availability markers as lessons", not fake_lessons, f"{len(fake_lessons)} fake availability lesson(s) detected.", {"fake_count": len(fake_lessons)}),
        gate("availability_sheet_does_not_create_lessons", "Availability sheets do not create lessons", all(sheet.get("scheduled_lessons_count", 0) == 0 for sheet in availability_sheets), "Availability-like sheets must have 0 scheduled lessons.", {"availability_sheets": len(availability_sheets)}),
        gate("no_requirements_without_hours", "No requirements without hours", not any("hours_not_parseable" in fake["reasons"] for fake in fake_requirements), "Requirements must have parseable hours.", {"fake_requirements": len(fake_requirements)}),
        gate("no_requirements_missing_core_fields", "No requirements missing core fields", not any({"missing_class", "missing_or_marker_subject", "missing_teacher"} & set(fake["reasons"]) for fake in fake_requirements), "Requirements must include class, subject and teacher.", {"fake_requirements": len(fake_requirements)}),
        gate("constraints_do_not_become_requirements", "Constraints do not become requirements", all(sheet.get("requirements_count", 0) == 0 and not sheet.get("fake_requirements") for sheet in constraints), "Constraint sheets must not produce requirements.", {"constraint_sheets": len(constraints)}),
        gate("mixed_lists_do_not_become_requirements", "Mixed lists do not become requirements", all(sheet.get("requirements_count", 0) == 0 and not sheet.get("fake_requirements") for sheet in mixed), "Mixed list sheets must not produce requirements.", {"mixed_list_sheets": len(mixed)}),
        gate("requirements_true_positive_stable", "Requirements true positive stable", any(sheet.get("actual_format") == "requirements_table" and sheet.get("requirements_count", 0) > 0 for sheet in requirement_sheets), "Known requirements sheet remains requirements_table and extracts >0 requirements.", {"requirement_sheets": len(requirement_sheets)}),
        gate("schedule_true_positive_stable", "Schedule true positive stable", any(sheet.get("actual_format") == "schedule_grid" and sheet.get("scheduled_lessons_count", 0) > 0 for sheet in schedule_sheets), "Known schedule sheet remains schedule_grid and extracts >0 lessons.", {"schedule_sheets": len(schedule_sheets)}),
        gate("noise_no_aggressive_extraction", "Noise no aggressive extraction", all(sheet.get("requirements_count", 0) == 0 and sheet.get("scheduled_lessons_count", 0) == 0 for sheet in noise_sheets), "Noise sheets must not produce requirements or lessons.", {"noise_sheets": len(noise_sheets)}),
        gate("v2_engine_used_in_eval", "v2 engine used in eval", bool(executed) and all(case.get("engine_used") == engine for case in executed), f"All executed cases must use {engine}.", {"engines": sorted({case.get("engine_used") for case in executed})}),
        gate("v1_default_preserved", "v1 default preserved", v1_default_ok, "Default analysis without EXCEL_INTELLIGENCE_MODE must remain v1.", {}),
        gate("low_real_file_coverage", "Low real file coverage", len([case for case in executed if not case.get("synthetic")]) >= 3, "Fewer than 3 real files were evaluated.", severity="warning"),
    ]


def build_report(engine: str, cases: list[dict[str, Any]], gates: list[dict[str, Any]], readiness: dict[str, Any], files: list[dict[str, Any]], *, fail_on_warning: bool = False) -> dict[str, Any]:
    checks = [check for case in cases for check in case.get("checks", [])]
    failures = [check for check in checks if not check.get("passed") and check.get("severity") == "blocking"]
    warnings = [check for check in checks if not check.get("passed") and check.get("severity") == "warning"]
    gate_failures = [gate for gate in gates if not gate["passed"] and gate["severity"] == "blocking"]
    gate_warnings = [gate for gate in gates if not gate["passed"] and gate["severity"] == "warning"]
    failed_warning_status = fail_on_warning and (warnings or gate_warnings)
    status = "failed" if failures or gate_failures or failed_warning_status else "passed"
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "engine": engine,
        "status": status,
        "summary": {
            "total_cases": len(cases),
            "total_checks": len(checks),
            "passed_checks": sum(1 for check in checks if check["passed"]),
            "failed_checks": sum(1 for check in checks if not check["passed"]),
            "warning_checks": len(warnings) + len(gate_warnings),
            "skipped_cases": sum(1 for case in cases if case["status"] == "skipped"),
            "blocking_failures": len(failures) + len(gate_failures),
            "readiness_score": readiness["score"],
            "recommendation": readiness["recommendation"],
        },
        "readiness": readiness,
        "gates": gates,
        "cases": cases,
        "failures": failures + gate_failures,
        "warnings": warnings + gate_warnings,
        "files": files,
    }


def inventory_files(cases: list[ExpectedCase]) -> list[dict[str, Any]]:
    items = []
    for case in cases:
        path = case.resolved_file_path(ROOT)
        exists = path.exists()
        items.append(
            {
                "case_id": case.case_id,
                "file_path": case.file_path,
                "exists": exists,
                "file_size": path.stat().st_size if exists else 0,
                "included_in_eval": exists or case.synthetic,
                "reason": "present" if exists else "synthetic_generated" if case.synthetic else "missing_skipped" if case.skip_if_file_missing else "missing_failure",
            }
        )
    return items


def create_synthetic_core_workbook(path: Path) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheets = {
        "synthetic_requirements": [["class", "subject", "teacher", "hours"], ["7A", "Math", "Cohen", 4], ["8B", "English", "Levi", "3 hours"]],
        "synthetic_schedule": [["שיעור", "ראשון", "שני"], ["08:00-08:45", "Math 7A\nמורה: Cohen\nחדר: 1", "English 8B\nמורה: Levi\nחדר: 2"]],
        "synthetic_availability": [["teacher", "Monday", "Tuesday"], ["Cohen", "available", "unavailable"], ["Levi", "yes", "no"]],
        "synthetic_noise": [["WhatsApp notes"], ["Reminder only, no import"], ["TODO: call teacher"]],
    }
    first = True
    for name, rows in sheets.items():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = name
        for row in rows:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    path.write_bytes(output.getvalue())


def _v1_default_preserved() -> bool:
    old = os.environ.pop("EXCEL_INTELLIGENCE_MODE", None)
    try:
        content = _tiny_requirements_workbook()
        result = analyze_excel_content(content, filename="v1-default-check.xlsx")
        return result.get("engine_used") == "v1"
    finally:
        if old is not None:
            os.environ["EXCEL_INTELLIGENCE_MODE"] = old


def _tiny_requirements_workbook() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["class", "subject", "teacher", "hours"])
    sheet.append(["7A", "Math", "Cohen", 4])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@contextmanager
def temporary_excel_mode(engine: str):
    old = os.environ.get("EXCEL_INTELLIGENCE_MODE")
    os.environ["EXCEL_INTELLIGENCE_MODE"] = engine
    try:
        yield
    finally:
        if old is None:
            os.environ.pop("EXCEL_INTELLIGENCE_MODE", None)
        else:
            os.environ["EXCEL_INTELLIGENCE_MODE"] = old


if __name__ == "__main__":
    raise SystemExit(main())
