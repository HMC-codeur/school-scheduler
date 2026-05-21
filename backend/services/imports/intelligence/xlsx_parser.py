from __future__ import annotations

from io import BytesIO

from backend.services.imports.intelligence.diagnostics import diagnostic
from backend.services.imports.intelligence.models import BrainResult, ImportContext, ImportSheet
from backend.services.imports.intelligence.normalizers import normalize_text


class XlsxParserBrain:
    name = "xlsx_parser"

    def run(self, context: ImportContext) -> BrainResult:
        if context.file_type not in {"xlsx", "xls"}:
            return context.add_result(BrainResult(self.name, "ok", 1.0, data={"skipped": True}))
        diagnostics = []
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(context.content), read_only=False, data_only=True)
        except Exception:
            diagnostics.append(
                diagnostic(
                    "excel_read_failed",
                    "blocking",
                    "Impossible de lire ce fichier Excel. Il semble vide, corrompu ou non supporté.",
                    suggestion="Réenregistrez le fichier en .xlsx.",
                    confidence=0.9,
                )
            )
            return context.add_result(BrainResult(self.name, "error", 0.0, diagnostics))

        try:
            context.sheets = [_sheet_from_openpyxl(sheet) for sheet in workbook.worksheets]
        finally:
            workbook.close()
        context.metadata["reader_used"] = "openpyxl"
        if not context.sheets:
            diagnostics.append(diagnostic("empty_workbook", "blocking", "Le fichier ne contient aucune feuille lisible."))
        return context.add_result(
            BrainResult(
                self.name,
                "error" if any(item.severity == "blocking" for item in diagnostics) else "ok",
                0.95 if context.sheets else 0.0,
                diagnostics,
                {"sheets_count": len(context.sheets), "reader_used": context.metadata.get("reader_used")},
            )
        )


def _sheet_from_openpyxl(sheet: object) -> ImportSheet:
    rows: dict[int, dict[int, str]] = {}
    max_row = int(getattr(sheet, "max_row", 0) or 0)
    max_column = int(getattr(sheet, "max_column", 0) or 0)
    for row in sheet.iter_rows():
        for cell in row:
            value = normalize_text(cell.value)
            if value:
                rows.setdefault(int(cell.row), {})[int(cell.column)] = value
    merged_ranges = getattr(getattr(sheet, "merged_cells", None), "ranges", [])
    return ImportSheet(
        name=str(getattr(sheet, "title", "Sheet")),
        rows=rows,
        max_row=max_row,
        max_column=max_column,
        merged_cells_count=len(list(merged_ranges)),
    )
