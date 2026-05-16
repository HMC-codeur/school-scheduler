from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, timedelta
import hashlib
import os
import re
import time
import zipfile
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any
from xml.etree import ElementTree as ET

from backend.models.schemas import (
    CommitResponse,
    ExcelImportCommitRequest,
    ExcelImportPreviewResponse,
    ImportedLesson,
    ImportError as ImportIssue,
    ImportWarning,
    ScheduleCell,
)
from backend.services.scoring import analyze_schedule, build_schedule_option
from backend.services.solver.stability import schedule_with_session_ids


NS_MAIN = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_RELS = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
PREVIEW_CACHE_TTL_SECONDS = 15 * 60
_PREVIEW_CACHE: dict[str, tuple[float, ExcelImportPreviewResponse]] = {}

DAY_ALIASES = {
    "mon": ("mon", "monday", "lun", "lundi", "שני", "יום שני"),
    "tue": ("tue", "tuesday", "mar", "mardi", "שלישי", "יום שלישי"),
    "wed": ("wed", "wednesday", "mer", "mercredi", "רביעי", "יום רביעי"),
    "thu": ("thu", "thursday", "jeu", "jeudi", "חמישי", "יום חמישי"),
    "fri": ("fri", "friday", "ven", "vendredi", "שישי", "יום שישי"),
    "sat": ("sat", "saturday", "sam", "samedi", "שבת", "יום שבת"),
    "sun": ("sun", "sunday", "dim", "dimanche", "ראשון", "יום ראשון"),
}
LABELS = {
    "teacher": ["מורה", "teacher", "prof", "professeur"],
    "room": ["חדר", "room", "salle"],
    "class_name": ["כיתה", "class", "classe"],
}


def excel_import_max_bytes() -> int:
    raw = os.getenv("EXCEL_IMPORT_MAX_BYTES", str(8 * 1024 * 1024))
    try:
        return max(1, int(raw))
    except ValueError:
        return 8 * 1024 * 1024


def preview_excel_schedule(
    content: bytes,
    filename: str | None = None,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    if len(content or b"") > excel_import_max_bytes():
        return _preview_response(
            filename=filename,
            errors=[_issue("file_too_large", f"Fichier trop volumineux: limite {excel_import_max_bytes()} octets.")],
        ).model_dump()
    warnings: list[ImportWarning] = []
    if filename and not filename.lower().endswith(".xlsx"):
        warnings.append(_warning("unsupported_extension", "Le fichier ne porte pas l'extension .xlsx.", value=filename))
    if not content:
        return _preview_response(filename=filename, warnings=warnings, errors=[_issue("empty_file", "Fichier vide.")]).model_dump()

    try:
        rows, selected_sheet, parser = _read_workbook_rows(content, sheet_name)
    except ValueError as exc:
        return _preview_response(filename=filename, warnings=warnings, errors=[_issue("unsupported_format", str(exc))]).model_dump()
    except zipfile.BadZipFile:
        return _preview_response(filename=filename, warnings=warnings, errors=[_issue("corrupt_file", "Fichier .xlsx invalide ou corrompu.")]).model_dump()
    except ET.ParseError:
        return _preview_response(filename=filename, warnings=warnings, errors=[_issue("corrupt_xml", "XML Excel invalide dans le fichier .xlsx.")]).model_dump()
    except KeyError as exc:
        return _preview_response(filename=filename, warnings=warnings, errors=[_issue("invalid_structure", f"Structure .xlsx incomplète : {exc}.")]).model_dump()
    except Exception as exc:
        return _preview_response(filename=filename, warnings=warnings, errors=[_issue("read_failed", f"Lecture Excel impossible: {exc}")]).model_dump()

    preview = _parse_grid_rows(rows, filename, selected_sheet, parser, warnings)
    if preview.can_commit:
        _cache_preview(preview)
    return preview.model_dump()


def commit_excel_import(payload: ExcelImportCommitRequest, store: Any) -> CommitResponse:
    lessons = payload.lessons
    if lessons is None and payload.import_id:
        preview = _get_cached_preview(payload.import_id)
        if preview is None:
            return _commit_error(payload, "Preview expirée ou inconnue. Relancez la preview ou envoyez lessons dans le body.")
        lessons = preview.lessons
    lessons = lessons or []
    if not lessons:
        return _commit_error(payload, "Aucune leçon exploitable à importer.")

    schedule, warnings, conflict_errors = _lessons_to_schedule(lessons)
    missing_errors = _unknown_entity_errors(lessons, store) if not payload.create_missing_entities else []
    errors = [*conflict_errors, *missing_errors]
    if errors and (payload.fail_on_conflict or missing_errors):
        return _commit_error(payload, "Import Excel refusé.", errors=errors, schedule=schedule)

    previous_schedule = store.schedule
    merged_schedule = _merge_schedule(previous_schedule, schedule, fail_on_conflict=payload.fail_on_conflict)
    if payload.mode == "merge":
        if merged_schedule["errors"]:
            return _commit_error(payload, "Import Excel refusé: conflit en mode merge.", errors=merged_schedule["errors"], schedule=previous_schedule)
        target_schedule = merged_schedule["schedule"]
    else:
        target_schedule = schedule

    created = {"classes": 0, "teachers": 0, "subjects": 0, "slots": 0}
    updated = {"teachers_subjects_augmented": 0}
    if not payload.dry_run and payload.create_missing_entities:
        created = _create_missing_entities(lessons, store)

    normalized_schedule = schedule_with_session_ids(target_schedule)
    diagnostics = active_schedule_diagnostics(
        normalized_schedule,
        store.classes,
        store.teachers,
        store.subjects,
        _slot_values(store.slots, normalized_schedule),
    )
    option_id = f"import-{_schedule_hash(normalized_schedule)}" if payload.synthesize_schedule_option else None

    if not payload.dry_run:
        store.schedule = normalized_schedule
        if option_id and payload.selected:
            option = build_schedule_option(
                option_id=option_id,
                schedule=normalized_schedule,
                classes=store.classes,
                teachers=store.teachers,
                subjects=store.subjects,
                slots=store.slots,
                constraints=store.conditions,
                learning_groups=getattr(store, "learning_groups", []),
            )
            option["id"] = option_id
            option["title"] = "Import Excel"
            option["selected"] = True
            option["message"] = "Planning importé depuis Excel."
            existing = [item for item in store.schedule_options if item.get("id") != option_id]
            store.schedule_options = [option, *existing]
            store.selected_schedule_option_id = option_id

    active_schedule = normalized_schedule if payload.dry_run else store.schedule
    return CommitResponse(
        success=True,
        message="Import Excel simulé avec succès." if payload.dry_run else "Import Excel appliqué avec succès.",
        mode=payload.mode,
        dry_run=payload.dry_run,
        warnings=warnings,
        errors=[],
        created_entities=created if not payload.dry_run else {"classes": 0, "teachers": 0, "subjects": 0, "slots": 0},
        updated_entities=updated,
        imported_lessons_count=len(lessons),
        active_schedule_entries_count=_schedule_size(active_schedule),
        schedule_option_id=option_id,
        selected_schedule_option_id=None if payload.dry_run else store.selected_schedule_option_id,
        schedule=active_schedule,
        diagnostics={"active_schedule": diagnostics},
        export_ready=bool(active_schedule),
        repair_ready=bool(active_schedule),
    )


def active_schedule_diagnostics(schedule: dict, classes: list, teachers: list, subjects: list, slots: list[str]) -> dict[str, Any]:
    metrics = analyze_schedule(schedule, classes, teachers, subjects, slots)
    normalized = _schedule_to_plain(schedule)
    duplicate_entries = 0
    seen: set[tuple[str, str, str, str]] = set()
    unknown = {"classes": [], "teachers": [], "subjects": [], "slots": []}
    class_names = {_norm(item.name): item.name for item in classes}
    teacher_names = {_norm(item.name): item.name for item in teachers}
    subject_names = {_norm(item.name): item.name for item in subjects}
    slot_names = set(slots)
    for slot, entries in normalized.items():
        if slot not in slot_names:
            unknown["slots"].append(slot)
        for class_name, cell in entries.items():
            key = (slot, _norm(class_name), _norm(cell.get("subject")), _norm(cell.get("teacher")))
            if key in seen:
                duplicate_entries += 1
            seen.add(key)
            if _norm(class_name) not in class_names:
                unknown["classes"].append(class_name)
            if _norm(cell.get("teacher")) not in teacher_names:
                unknown["teachers"].append(cell.get("teacher", ""))
            if _norm(cell.get("subject")) not in subject_names:
                unknown["subjects"].append(cell.get("subject", ""))
    return {
        "teacher_conflicts": int(metrics.get("teacher_conflicts", 0)),
        "class_conflicts": int(metrics.get("class_conflicts", 0)),
        "duplicate_entries": duplicate_entries,
        "unknown_entities": {key: _unique_preserve_order(values) for key, values in unknown.items()},
        "empty_gaps": int(metrics.get("empty_gaps", 0)),
        "metrics": metrics,
    }


def _read_workbook_rows(content: bytes, sheet_name: str | None) -> tuple[dict[int, dict[int, str]], str | None, str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        rows = _read_sheet_rows_xml(content, sheet_name)
        return rows["rows"], rows["sheet_name"], "xml-fallback"

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        selected = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise ValueError(f"Feuille introuvable: {sheet_name}")
        sheet = workbook[selected]
        rows: dict[int, dict[int, str]] = {}
        for row in sheet.iter_rows():
            for cell in row:
                value = _normalize_text(cell.value)
                if value:
                    rows.setdefault(cell.row, {})[cell.column] = value
        return rows, selected, "openpyxl"
    finally:
        workbook.close()


def _parse_grid_rows(
    rows: dict[int, dict[int, str]],
    filename: str | None,
    sheet_name: str | None,
    parser: str,
    warnings: list[ImportWarning],
) -> ExcelImportPreviewResponse:
    if not rows:
        return _preview_response(filename=filename, sheet_name=sheet_name, parser_used=parser, warnings=warnings, errors=[_issue("empty_sheet", "Feuille vide.")])
    header_index = _find_header_row(rows)
    if header_index is None:
        return _preview_response(filename=filename, sheet_name=sheet_name, parser_used=parser, warnings=warnings, errors=[_issue("missing_days", "Impossible de détecter la ligne des jours.")])
    header_row = rows[header_index]
    day_columns = [(column, value) for column, value in sorted(header_row.items()) if column > 1 and value]
    day_columns = [(column, value) for column, value in day_columns if _day_key(value)]
    if not day_columns:
        return _preview_response(filename=filename, sheet_name=sheet_name, parser_used=parser, warnings=warnings, errors=[_issue("missing_days", "Aucune colonne de jour détectée.")])

    lessons: list[ImportedLesson] = []
    slots: list[str] = []
    first_day_column = min(column for column, _ in day_columns)
    for row_number in sorted(row for row in rows if row > header_index):
        row = rows[row_number]
        slot_label, start_time, end_time = _parse_slot(_detect_slot(row, first_day_column))
        if not slot_label:
            if any(_normalize_text(row.get(column, "")) for column, _ in day_columns):
                warnings.append(_warning("missing_slot", f"Ligne {row_number} ignorée: créneau absent.", row=row_number))
            continue
        slots.append(slot_label)
        for column, day in day_columns:
            raw = _normalize_text(row.get(column, ""))
            if not raw:
                continue
            parsed = _parse_lesson_cell(raw)
            day_key = _day_key(day) or _slug(day)
            slot_key = f"{day_key.title()}-{start_time}" if start_time else None
            lesson_warnings = _lesson_warnings(parsed, row_number, column, raw)
            lessons.append(
                ImportedLesson(
                    day=day,
                    day_key=day_key,
                    slot=slot_label,
                    slot_label=slot_label,
                    slot_key=slot_key,
                    start_time=start_time,
                    end_time=end_time,
                    row=row_number,
                    column=column,
                    raw=raw,
                    session_id=_session_id(day_key, start_time, parsed.get("class_name"), parsed.get("subject"), len(lessons) + 1),
                    normalized={key: _norm(value) for key, value in parsed.items() if value},
                    warnings=lesson_warnings,
                    **parsed,
                )
            )
            warnings.extend(lesson_warnings)
    if not lessons:
        return _preview_response(filename=filename, sheet_name=sheet_name, parser_used=parser, warnings=warnings, errors=[_issue("no_lessons", "Aucune leçon exploitable détectée.")])
    preview = _preview_response(filename=filename, sheet_name=sheet_name, parser_used=parser, warnings=warnings, lessons=lessons)
    return preview


def _preview_response(
    *,
    filename: str | None = None,
    lessons: list[ImportedLesson] | None = None,
    warnings: list[ImportWarning] | None = None,
    errors: list[ImportIssue] | None = None,
    sheet_name: str | None = None,
    parser_used: str | None = None,
) -> ExcelImportPreviewResponse:
    lessons = lessons or []
    warnings = warnings or []
    errors = errors or []
    preview_hash = _lessons_hash(lessons) if lessons else None
    import_id = f"imp_{preview_hash}" if preview_hash and not errors else None
    days = _unique_preserve_order([lesson.day for lesson in lessons])
    slots = _unique_preserve_order([lesson.slot_label or lesson.slot for lesson in lessons])
    classes = _unique_preserve_order([lesson.class_name or "" for lesson in lessons])
    teachers = _unique_preserve_order([lesson.teacher or "" for lesson in lessons])
    subjects = _unique_preserve_order([lesson.subject or "" for lesson in lessons])
    rooms = _unique_preserve_order([lesson.room or "" for lesson in lessons])
    return ExcelImportPreviewResponse(
        filename=filename,
        days=days,
        slots=slots,
        classes=classes,
        teachers=teachers,
        subjects=subjects,
        rooms=rooms,
        lessons=lessons,
        counts={"days": len(days), "slots": len(slots), "classes": len(classes), "teachers": len(teachers), "subjects": len(subjects), "rooms": len(rooms), "lessons": len(lessons)},
        warnings=_unique_preserve_order([warning.message for warning in warnings]),
        errors=_unique_preserve_order([error.message for error in errors]),
        warning_details=warnings,
        error_details=errors,
        can_commit=bool(lessons and not errors),
        import_id=import_id,
        preview_hash=preview_hash,
        sheet_name=sheet_name,
        parser_used=parser_used,
    )


def _lessons_to_schedule(lessons: list[ImportedLesson]) -> tuple[dict[str, dict[str, ScheduleCell]], list[str], list[str]]:
    schedule: dict[str, dict[str, ScheduleCell]] = {}
    warnings: list[str] = []
    errors: list[str] = []
    teacher_slot: Counter[tuple[str, str]] = Counter()
    class_slot: Counter[tuple[str, str]] = Counter()
    exact: set[tuple[str, str, str, str]] = set()
    for index, lesson in enumerate(lessons):
        if not lesson.slot_key:
            errors.append(f"Leçon {index}: créneau absent ou illisible.")
            continue
        if not lesson.class_name:
            errors.append(f"Leçon {index}: classe absente.")
            continue
        if not lesson.subject:
            errors.append(f"Leçon {index}: matière absente.")
            continue
        if not lesson.teacher:
            warnings.append(f"Leçon {index}: professeur absent.")
        teacher = lesson.teacher or ""
        key = (lesson.slot_key, lesson.class_name, lesson.subject, teacher)
        if key in exact:
            errors.append(f"Doublon exact: {lesson.slot_key} / {lesson.class_name} / {lesson.subject}.")
        exact.add(key)
        teacher_slot[(lesson.slot_key, _norm(teacher))] += 1
        class_slot[(lesson.slot_key, _norm(lesson.class_name))] += 1
        schedule.setdefault(lesson.slot_key, {})[lesson.class_name] = ScheduleCell(
            subject=lesson.subject,
            teacher=teacher,
            session_id=lesson.session_id or _session_id(lesson.day_key, lesson.start_time, lesson.class_name, lesson.subject, index + 1),
        )
    for (slot, teacher), count in teacher_slot.items():
        if teacher and count > 1:
            errors.append(f"Conflit professeur: {teacher} a {count} cours sur {slot}.")
    for (slot, class_name), count in class_slot.items():
        if class_name and count > 1:
            errors.append(f"Conflit classe: {class_name} a {count} cours sur {slot}.")
    return schedule, _unique_preserve_order(warnings), _unique_preserve_order(errors)


def _unknown_entity_errors(lessons: list[ImportedLesson], store: Any) -> list[str]:
    existing = {
        "classes": {_norm(item.name) for item in store.classes},
        "teachers": {_norm(item.name) for item in store.teachers},
        "subjects": {_norm(item.name) for item in store.subjects},
        "slots": {_norm(slot) for slot in store.slots},
    }
    missing = {
        "classes": [lesson.class_name for lesson in lessons if lesson.class_name and _norm(lesson.class_name) not in existing["classes"]],
        "teachers": [lesson.teacher for lesson in lessons if lesson.teacher and _norm(lesson.teacher) not in existing["teachers"]],
        "subjects": [lesson.subject for lesson in lessons if lesson.subject and _norm(lesson.subject) not in existing["subjects"]],
        "slots": [lesson.slot_key for lesson in lessons if lesson.slot_key and _norm(lesson.slot_key) not in existing["slots"]],
    }
    errors = []
    for key, values in missing.items():
        unique = _unique_preserve_order([value or "" for value in values])
        if unique:
            errors.append(f"Entités inconnues ({key}): {', '.join(unique)}.")
    return errors


def _create_missing_entities(lessons: list[ImportedLesson], store: Any) -> dict[str, int]:
    created = {"classes": 0, "teachers": 0, "subjects": 0, "slots": 0}
    classes = {_norm(item.name) for item in store.classes}
    subjects = {_norm(item.name) for item in store.subjects}
    teachers = {_norm(item.name) for item in store.teachers}
    slots = {_norm(slot) for slot in store.slots}
    for class_name in _unique_preserve_order([lesson.class_name or "" for lesson in lessons]):
        if _norm(class_name) not in classes:
            store.add_class(class_name)
            classes.add(_norm(class_name))
            created["classes"] += 1
    for subject in _unique_preserve_order([lesson.subject or "" for lesson in lessons]):
        if _norm(subject) not in subjects:
            store.add_subject(subject, 1)
            subjects.add(_norm(subject))
            created["subjects"] += 1
    teacher_subjects: dict[str, set[str]] = defaultdict(set)
    teacher_display: dict[str, str] = {}
    for lesson in lessons:
        if lesson.teacher:
            teacher_display[_norm(lesson.teacher)] = lesson.teacher
            if lesson.subject:
                teacher_subjects[_norm(lesson.teacher)].add(lesson.subject)
    for teacher_norm, subject_set in teacher_subjects.items():
        if teacher_norm not in teachers:
            store.add_teacher(teacher_display[teacher_norm], sorted(subject_set))
            teachers.add(teacher_norm)
            created["teachers"] += 1
    for slot in _unique_preserve_order([lesson.slot_key or "" for lesson in lessons]):
        if _norm(slot) not in slots:
            store.add_slot(slot)
            slots.add(_norm(slot))
            created["slots"] += 1
    return created


def _merge_schedule(current: dict, incoming: dict, *, fail_on_conflict: bool) -> dict[str, Any]:
    merged = _schedule_to_plain(current)
    errors: list[str] = []
    for slot, entries in _schedule_to_plain(incoming).items():
        merged.setdefault(slot, {})
        existing_teachers = {cell.get("teacher") for cell in merged[slot].values() if cell.get("teacher")}
        for class_name, cell in entries.items():
            if class_name in merged[slot]:
                errors.append(f"Conflit merge: {slot} / {class_name} existe déjà.")
            if cell.get("teacher") and cell.get("teacher") in existing_teachers:
                errors.append(f"Conflit merge professeur: {cell.get('teacher')} sur {slot}.")
            if not errors or not fail_on_conflict:
                merged[slot][class_name] = ScheduleCell(**cell)
    return {"schedule": merged, "errors": _unique_preserve_order(errors)}


def _commit_error(payload: ExcelImportCommitRequest, message: str, *, errors: list[str] | None = None, schedule: dict | None = None) -> CommitResponse:
    return CommitResponse(
        success=False,
        message=message,
        mode=payload.mode,
        dry_run=payload.dry_run,
        errors=errors or [message],
        schedule=schedule or {},
        diagnostics={"active_schedule": active_schedule_diagnostics(schedule or {}, [], [], [], [])},
        export_ready=False,
        repair_ready=False,
    )


def _read_sheet_rows_xml(content: bytes, sheet_name: str | None = None) -> dict[str, Any]:
    with zipfile.ZipFile(BytesIO(content)) as archive:
        shared_strings = _read_shared_strings(archive)
        sheet_path, selected_name = _sheet_path(archive, sheet_name)
        worksheet = ET.fromstring(archive.read(sheet_path))
    rows: dict[int, dict[int, str]] = {}
    for cell in worksheet.findall(".//x:sheetData/x:row/x:c", NS_MAIN):
        row_number, column_number = _cell_coordinates(cell.attrib.get("r", ""))
        if row_number is None or column_number is None:
            continue
        value = _cell_text(cell, shared_strings)
        if value:
            rows.setdefault(row_number, {})[column_number] = value
    return {"rows": rows, "sheet_name": selected_name}


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    return [_normalize_text("".join(text.text or "" for text in item.findall(".//x:t", NS_MAIN))) for item in root.findall("x:si", NS_MAIN)]


def _sheet_path(archive: zipfile.ZipFile, sheet_name: str | None) -> tuple[str, str | None]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    sheets = workbook.findall(".//x:sheets/x:sheet", NS_MAIN)
    if not sheets:
        raise KeyError("xl/workbook.xml sans feuille")
    selected = next((sheet for sheet in sheets if sheet.attrib.get("name") == sheet_name), None) if sheet_name else sheets[0]
    if selected is None:
        raise ValueError(f"Feuille introuvable: {sheet_name}")
    selected_name = selected.attrib.get("name")
    rel_id = selected.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
    if not rel_id:
        return "xl/worksheets/sheet1.xml", selected_name
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for rel in rels.findall("r:Relationship", NS_RELS):
        if rel.attrib.get("Id") == rel_id:
            target = rel.attrib.get("Target", "worksheets/sheet1.xml")
            return str(PurePosixPath("xl") / target).replace("\\", "/"), selected_name
    return "xl/worksheets/sheet1.xml", selected_name


def _cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return _normalize_text("".join(text.text or "" for text in cell.findall(".//x:t", NS_MAIN)))
    value = cell.find("x:v", NS_MAIN)
    if value is None or value.text is None:
        return ""
    if cell_type == "s":
        try:
            return shared_strings[int(value.text)]
        except (ValueError, IndexError):
            return ""
    return _normalize_text(value.text)


def _cell_coordinates(reference: str) -> tuple[int | None, int | None]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", reference or "")
    if not match:
        return None, None
    column = 0
    for char in match.group(1):
        column = column * 26 + (ord(char) - ord("A") + 1)
    return int(match.group(2)), column


def _find_header_row(rows: dict[int, dict[int, str]]) -> int | None:
    best_row: int | None = None
    best_score = 0
    for row_number, row in rows.items():
        values = [_normalize_text(value) for column, value in sorted(row.items()) if column > 1 and value]
        score = sum(1 for value in values if _day_key(value))
        if score > best_score:
            best_score = score
            best_row = row_number
    return best_row if best_score > 0 else None


def _detect_slot(row: dict[int, str], first_day_column: int) -> str:
    for column in range(1, first_day_column):
        value = _normalize_text(row.get(column, ""))
        if value:
            return value
    return _normalize_text(row.get(1, ""))


def _parse_slot(value: str) -> tuple[str | None, str | None, str | None]:
    cleaned = _normalize_text(value)
    if not cleaned:
        return None, None, None
    times = re.findall(r"([01]?\d|2[0-3])[:hH]([0-5]\d)", cleaned)
    if not times:
        return cleaned, None, None
    normalized = [f"{int(hour):02d}:{minute}" for hour, minute in times]
    start = normalized[0]
    end = normalized[1] if len(normalized) > 1 else None
    return (f"{start}-{end}" if end else start), start, end


def _parse_lesson_cell(raw: str) -> dict[str, str | None]:
    teacher = _extract_label_value(raw, LABELS["teacher"])
    room = _extract_label_value(raw, LABELS["room"])
    explicit_class = _extract_label_value(raw, LABELS["class_name"])
    content_lines = []
    for line in raw.split("\n"):
        if _line_has_label(line, [*LABELS["teacher"], *LABELS["room"], *LABELS["class_name"]]):
            continue
        cleaned = _normalize_text(line)
        if cleaned:
            content_lines.append(cleaned)
    subject_line = content_lines[0] if content_lines else ""
    class_name = explicit_class or _extract_class_name(subject_line) or _extract_class_name(raw)
    subject = _subject_from_line(subject_line, class_name)
    return {"subject": subject or None, "class_name": class_name, "teacher": teacher, "room": room}


def _extract_label_value(text: str, labels: list[str]) -> str | None:
    joined = "\n".join(_normalize_text(line) for line in text.split("\n"))
    for label in labels:
        match = re.search(rf"(?:^|\n)\s*{re.escape(label)}\s*[:：]\s*([^\n]+)", joined, flags=re.IGNORECASE)
        if match:
            return _normalize_text(match.group(1)) or None
    return None


def _line_has_label(line: str, labels: list[str]) -> bool:
    return any(re.search(rf"^\s*{re.escape(label)}\s*[:：]", line, flags=re.IGNORECASE) for label in labels)


def _extract_class_name(text: str) -> str | None:
    explicit = re.search(r"(?:כיתה|classe|class)\s*[:：-]?\s*([0-9A-Za-zא-ת\"'׳ -]{1,16})", text, flags=re.IGNORECASE)
    if explicit:
        return _clean_class_token(explicit.group(1))
    tokens = [token.strip("()[]{}.,;:") for token in text.split()]
    for token in reversed(tokens):
        cleaned = _clean_class_token(token)
        if cleaned and _looks_like_class_token(cleaned):
            return cleaned
    return None


def _clean_class_token(value: str) -> str | None:
    cleaned = _normalize_text(value).strip("()[]{}.,;:")
    return cleaned or None


def _looks_like_class_token(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,2}[A-Za-zא-ת]?", value) or re.fullmatch(r"[A-Za-zא-ת]\d{1,2}", value) or re.fullmatch(r"[זחטיכל][\"'׳]?\d?", value))


def _subject_from_line(subject_line: str, class_name: str | None) -> str:
    subject = re.sub(r"(?:כיתה|classe|class)\s*[:：-]?\s*[0-9A-Za-zא-ת\"'׳ -]{1,16}", "", subject_line, flags=re.IGNORECASE).strip()
    if class_name:
        subject = re.sub(rf"(?:^|\s){re.escape(class_name)}(?:$|\s)", " ", subject).strip()
    return _normalize_text(subject)


def _lesson_warnings(parsed: dict[str, str | None], row: int, column: int, raw: str) -> list[ImportWarning]:
    warnings = []
    if not parsed.get("teacher"):
        warnings.append(_warning("missing_teacher", "Professeur absent ou indétectable.", row=row, column=column, value=raw))
    if not parsed.get("subject"):
        warnings.append(_warning("missing_subject", "Matière absente ou indétectable.", row=row, column=column, value=raw))
    if not parsed.get("class_name"):
        warnings.append(_warning("missing_class", "Classe absente ou indétectable.", row=row, column=column, value=raw))
    return warnings


def _day_key(value: str) -> str | None:
    cleaned = re.sub(r"^יום\s+", "", _normalize_text(value).lower())
    for key, aliases in DAY_ALIASES.items():
        if cleaned in aliases or any(alias in cleaned for alias in aliases if len(alias) > 3):
            return key
    return None


def _session_id(day_key: str | None, start_time: str | None, class_name: str | None, subject: str | None, index: int) -> str:
    base = "_".join([day_key or "day", (start_time or "time").replace(":", ""), _slug(class_name or "class"), _slug(subject or "subject")])
    return f"imp_{base}_{index:03d}"[:80]


def _schedule_to_plain(schedule: dict | None) -> dict[str, dict[str, dict[str, str | None]]]:
    plain: dict[str, dict[str, dict[str, str | None]]] = {}
    for slot, entries in (schedule or {}).items():
        plain[str(slot)] = {}
        for class_name, cell in (entries or {}).items():
            if isinstance(cell, ScheduleCell):
                plain[str(slot)][str(class_name)] = cell.model_dump()
            elif isinstance(cell, dict):
                plain[str(slot)][str(class_name)] = {"subject": str(cell.get("subject", "")), "teacher": str(cell.get("teacher", "")), "session_id": cell.get("session_id")}
    return plain


def _slot_values(slots: list[str], schedule: dict) -> list[str]:
    return _unique_preserve_order([*slots, *list((schedule or {}).keys())])


def _schedule_size(schedule: dict) -> int:
    return sum(len(entries or {}) for entries in (schedule or {}).values())


def _lessons_hash(lessons: list[ImportedLesson]) -> str:
    payload = [lesson.model_dump(mode="json") for lesson in lessons]
    return hashlib.sha1(repr(payload).encode("utf-8")).hexdigest()[:8]


def _schedule_hash(schedule: dict) -> str:
    return hashlib.sha1(repr(_schedule_to_plain(schedule)).encode("utf-8")).hexdigest()[:8]


def _cache_preview(preview: ExcelImportPreviewResponse) -> None:
    _purge_preview_cache()
    if preview.import_id:
        _PREVIEW_CACHE[preview.import_id] = (time.time(), preview)


def _get_cached_preview(import_id: str) -> ExcelImportPreviewResponse | None:
    _purge_preview_cache()
    cached = _PREVIEW_CACHE.get(import_id)
    return cached[1] if cached else None


def _purge_preview_cache() -> None:
    now = time.time()
    for key, (created, _) in list(_PREVIEW_CACHE.items()):
        if now - created > PREVIEW_CACHE_TTL_SECONDS:
            _PREVIEW_CACHE.pop(key, None)


def _warning(code: str, message: str, row: int | None = None, column: int | None = None, value: Any | None = None, lesson_index: int | None = None) -> ImportWarning:
    return ImportWarning(code=code, message=message, row=row, column=column, value=value, lesson_index=lesson_index)


def _issue(code: str, message: str, row: int | None = None, column: int | None = None, value: Any | None = None, lesson_index: int | None = None) -> ImportIssue:
    return ImportIssue(code=code, message=message, row=row, column=column, value=value, lesson_index=lesson_index)


def _normalize_text(value: Any) -> str:
    lines = [re.sub(r"\s+", " ", line).strip() for line in str(value or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    return "\n".join(line for line in lines if line)


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _slug(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-zא-ת]+", "_", _normalize_text(value)).strip("_")
    return cleaned or "x"


def _unique_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for value in values:
        cleaned = _normalize_text(value)
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        unique.append(cleaned)
    return unique
