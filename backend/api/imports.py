from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from io import BytesIO, StringIO
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile

from backend.data.repository import SchedulerRepository
from backend.data.store import get_store
from backend.models.schemas import ExcelImportCommitRequest
from backend.services.excel_import import commit_excel_import, excel_import_max_bytes, preview_excel_schedule
from backend.services.imports.intelligence.normalizers import day_key, is_time_like, looks_like_class_token, normalize_text, parse_lesson_cell
from backend.services.imports.intelligence.school_terms import looks_availability_like, looks_constraint_like, looks_noise_like


router = APIRouter(prefix="/imports", tags=["imports"])

_ANALYSES: dict[str, dict[str, Any]] = {}


@router.post("/analyze", response_model=dict)
async def analyze_import(file: UploadFile = File(...)) -> dict:
    content = await file.read()
    if len(content or b"") > excel_import_max_bytes():
        raise HTTPException(status_code=413, detail=f"Import file is too large. Limit is {excel_import_max_bytes()} bytes.")
    result = _analyze_import_primary(content, filename=file.filename)
    _ANALYSES[result["import_id"]] = result
    return result


@router.get("/{import_id}", response_model=dict)
def get_import(import_id: str) -> dict:
    result = _ANALYSES.get(import_id)
    if result is None:
        raise HTTPException(status_code=404, detail=f"Import analysis '{import_id}' not found")
    return result


@router.post("/{import_id}/confirm", response_model=dict)
def confirm_import(import_id: str, corrections: dict[str, Any] | None = None) -> dict:
    result = _ANALYSES.get(import_id)
    if result is None:
        raise HTTPException(status_code=404, detail=f"Import analysis '{import_id}' not found")
    confirmed = {**result, "status": "confirmed", "corrections": corrections or {}}
    _ANALYSES[import_id] = confirmed
    return confirmed


@router.post("/{import_id}/apply", response_model=dict)
def apply_import(import_id: str, store: SchedulerRepository = Depends(get_store)) -> dict:
    result = _ANALYSES.get(import_id)
    if result is None:
        raise HTTPException(status_code=404, detail=f"Import analysis '{import_id}' not found")
    preview = result.get("normalized_preview", {})
    created = _create_missing_entities(preview, store)
    return {
        "import_id": import_id,
        "status": "applied",
        "created_entities": created,
        "message": "Import analysis applied to reference data.",
    }


@router.post("/excel/analyze", response_model=dict)
async def analyze_excel(
    file: UploadFile | None = File(default=None),
    corrections: str | None = Form(default=None),
    debug_compare: bool = Query(default=False),
) -> dict:
    if file is None:
        raise HTTPException(status_code=400, detail="Aucun fichier Excel reçu.")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Aucun fichier Excel reçu.")
    if len(content or b"") > excel_import_max_bytes():
        raise HTTPException(status_code=413, detail=f"Excel file is too large. Limit is {excel_import_max_bytes()} bytes.")
    if not _is_excel_filename(file.filename):
        raise HTTPException(status_code=400, detail="Format invalide. Utilisez un fichier .xlsx ou .xlsm.")
    mvp_result = _try_excel_mvp_analyze(content, file.filename, corrections, debug_compare)
    if mvp_result is not None:
        return mvp_result
    preview = preview_excel_schedule(content, filename=file.filename)
    if preview.get("errors"):
        result = _analyze_xlsx_table(content, file.filename)
        result["engine_used"] = "v1"
        return result
    return _excel_preview_to_analysis(preview, filename=file.filename, legacy=True)


@router.post("/validate-grid-candidates", response_model=dict)
def validate_grid_candidates(payload: dict[str, Any] | list[Any]) -> dict:
    candidates = payload if isinstance(payload, list) else payload.get("candidates", [])
    if not isinstance(candidates, list):
        raise HTTPException(status_code=400, detail="candidates must be a list")

    valid_candidates: list[dict[str, Any]] = []
    rejected_candidates: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    blocking_errors: list[dict[str, Any]] = []
    ignored_count = 0
    accepted_count = 0

    for index, raw_candidate in enumerate(candidates):
        candidate = raw_candidate if isinstance(raw_candidate, dict) else {"raw": raw_candidate}
        status = normalize_text(candidate.get("status") or candidate.get("review_status") or "accepted").casefold()
        if status == "ignored":
            ignored_count += 1
            continue
        if status != "accepted":
            error = _grid_candidate_issue(index, candidate, "unsupported_status", "Statut de candidat non supporté.", field="status", value=status)
            rejected_candidates.append({"index": index, "candidate": candidate, "errors": [error], "suggestion": _grid_candidate_suggestion(candidate, [error])})
            blocking_errors.append(error)
            continue

        accepted_count += 1
        normalized, errors = _validate_accepted_grid_candidate(candidate, index)
        if errors:
            rejected_candidates.append({"index": index, "candidate": candidate, "errors": errors, "suggestion": _grid_candidate_suggestion(candidate, errors)})
            blocking_errors.extend(errors)
            continue

        confidence = normalized.get("confidence")
        if confidence is not None and confidence < 0.6:
            warnings.append(
                _grid_candidate_issue(
                    index,
                    candidate,
                    "low_confidence_accepted",
                    "Candidat accepté avec confiance faible.",
                    field="confidence",
                    value=confidence,
                )
            )
        valid_candidates.append(normalized)

    return {
        "valid_candidates": valid_candidates,
        "rejected_candidates": rejected_candidates,
        "warnings": warnings,
        "blocking_errors": blocking_errors,
        "summary": {
            "total_candidates": len(candidates),
            "accepted_candidates": accepted_count,
            "ignored_candidates": ignored_count,
            "valid_candidates": len(valid_candidates),
            "rejected_candidates": len(rejected_candidates),
            "warnings": len(warnings),
            "blocking_errors": len(blocking_errors),
        },
        "can_import": False,
        "requires_final_confirmation": True,
        "dry_run": True,
    }


@router.get("/excel/schema", response_model=dict)
def get_excel_schema() -> dict:
    return {
        "standard_fields": ["class_name", "subject_name", "teacher_name", "weekly_hours"],
        "supported_files": [".csv", ".xlsx"],
    }


@router.post("/excel/{import_id}/commit", response_model=dict)
def commit_excel(import_id: str, store: SchedulerRepository = Depends(get_store)) -> dict:
    mvp_result = _try_excel_mvp_commit(import_id, store)
    if mvp_result is not None:
        return mvp_result
    response = commit_excel_import(ExcelImportCommitRequest(import_id=import_id), store)
    if not response.success:
        raise HTTPException(status_code=400, detail=response.model_dump(mode="json"))
    return response.model_dump(mode="json")


def _try_excel_mvp_analyze(
    content: bytes,
    filename: str | None,
    corrections: str | None,
    debug_compare: bool,
) -> dict[str, Any] | None:
    try:
        from backend.services.imports.excel_mvp import analyze_excel_content, analyze_excel_content_debug_compare
        from backend.services.imports.excel_readers import ExcelReadError
    except (ImportError, ModuleNotFoundError):
        return None

    try:
        parsed_corrections = _parse_corrections(corrections)
        if debug_compare:
            return analyze_excel_content_debug_compare(content, filename=filename, corrections=parsed_corrections)
        return analyze_excel_content(content, filename=filename, corrections=parsed_corrections)
    except ExcelReadError as exc:
        raise HTTPException(status_code=400, detail=getattr(exc, "message", "Lecture Excel impossible.")) from exc
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Le fichier Excel semble vide, corrompu ou non supporté.") from exc


def _try_excel_mvp_commit(import_id: str, store: SchedulerRepository) -> dict[str, Any] | None:
    try:
        from backend.services.imports.excel_mvp import commit_excel_mvp_import, get_import_draft
    except (ImportError, ModuleNotFoundError):
        return None

    draft = get_import_draft(import_id)
    if draft is None:
        raise HTTPException(status_code=404, detail=f"Import draft '{import_id}' not found")
    result = commit_excel_mvp_import(draft, store)
    if result.get("status") == "blocked":
        raise HTTPException(status_code=409, detail=result)
    return result


def _validate_accepted_grid_candidate(candidate: dict[str, Any], index: int) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    class_name = _first_text(candidate, "class_name", "class_group", "group", "class")
    day = _first_text(candidate, "day")
    slot = _first_text(candidate, "slot", "time", "start_time", "period")
    subject = _first_text(candidate, "subject", "subject_name")
    teacher = _first_text(candidate, "teacher", "teacher_name")
    raw_cell = _first_text(candidate, "original_text", "raw_cell", "cell_text")
    confidence = _optional_float(candidate.get("confidence"))

    errors: list[dict[str, Any]] = []
    if not class_name:
        errors.append(_grid_candidate_issue(index, candidate, "missing_class_or_group", "Classe ou groupe obligatoire.", field="class_name"))
    if not _recognizable_day(day, slot):
        errors.append(_grid_candidate_issue(index, candidate, "missing_or_unrecognized_day", "Jour obligatoire ou non reconnu.", field="day", value=day))
    if not _recognizable_slot(slot):
        errors.append(_grid_candidate_issue(index, candidate, "missing_or_unrecognized_slot", "Créneau ou heure obligatoire ou non reconnu.", field="slot", value=slot))
    if not subject:
        errors.append(_grid_candidate_issue(index, candidate, "missing_subject", "Matière obligatoire.", field="subject"))

    normalized = {
        "class_name": class_name,
        "day": day,
        "day_key": day_key(day) or day_key(slot),
        "slot": slot,
        "subject": subject,
        "teacher": teacher or None,
        "raw_cell": raw_cell or None,
        "confidence": confidence,
        "source_trace": candidate.get("source_trace") if isinstance(candidate.get("source_trace"), dict) else None,
        "status": "accepted",
    }
    return normalized, errors


def _first_text(candidate: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = normalize_text(candidate.get(key))
        if value:
            return value
    return ""


def _recognizable_day(day: str, slot: str) -> bool:
    return bool(day_key(day) or day_key(slot))


def _recognizable_slot(slot: str) -> bool:
    return bool(slot and is_time_like(slot))


def _optional_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _grid_candidate_suggestion(candidate: dict[str, Any], errors: list[dict[str, Any]]) -> dict[str, Any]:
    raw_cell = _first_text(candidate, "original_text", "raw_cell", "cell_text", "text", "raw")
    error_codes = {str(error.get("code") or "") for error in errors}
    parsed, parse_warnings = parse_lesson_cell(raw_cell)
    proposed_class = _extract_class_token_from_text(raw_cell) or normalize_text(parsed.get("class_name"))
    proposed_subject = _subject_from_raw_cell(raw_cell, parsed)

    if "missing_class_or_group" in error_codes and proposed_class:
        return {
            "action": "fill_missing_class",
            "label_he": "השלמת כיתה",
            "explanation_he": f"נראה שחסרה כיתה בשורה הזו. הצעה: לשייך לכיתה {proposed_class}",
            "proposed_values": {"class_name": proposed_class},
            "confidence": 0.82,
        }
    if _looks_like_non_lesson(raw_cell):
        return {
            "action": "ignore_as_non_lesson",
            "label_he": "התעלמות משורה שאינה שיעור",
            "explanation_he": "השורה הזו לא נראית כמו שיעור ולכן מומלץ להתעלם ממנה.",
            "confidence": 0.86,
        }
    if "missing_subject" in error_codes and proposed_subject:
        confidence = 0.72 if not parse_warnings else 0.58
        return {
            "action": "edit_subject",
            "label_he": "השלמת מקצוע",
            "explanation_he": f"נראה שחסר מקצוע בשורה הזו. הצעה: {proposed_subject}",
            "proposed_values": {"subject": proposed_subject},
            "confidence": confidence,
        }
    if error_codes & {"missing_or_unrecognized_day", "missing_or_unrecognized_slot"}:
        return {
            "action": "edit_day_or_slot",
            "label_he": "תיקון יום או שעה",
            "explanation_he": "היום או השעה לא זוהו בוודאות. מומלץ לפתוח את השורה לעריכה ידנית.",
            "confidence": 0.62,
        }
    return {
        "action": "manual_review",
        "label_he": "בדיקה ידנית",
        "explanation_he": "נדרשת בדיקה ידנית.",
        "confidence": 0.5,
    }


def _looks_like_non_lesson(raw_cell: str) -> bool:
    return bool(raw_cell and (looks_availability_like(raw_cell) or looks_constraint_like(raw_cell) or looks_noise_like(raw_cell)))


def _extract_class_token_from_text(raw_cell: str) -> str | None:
    if not raw_cell:
        return None
    tokens = re.findall(r"[\w\u0590-\u05ff\"'׳]+", raw_cell)
    candidates: list[str] = []
    for token in tokens:
        parts = [token]
        parts.extend(part for part in re.split(r"[-/]", token) if part)
        candidates.extend(parts)
    for token in reversed(candidates):
        cleaned = token.strip("()[]{}.,;:")
        if looks_like_class_token(cleaned):
            return cleaned
    return None


def _subject_from_raw_cell(raw_cell: str, parsed: dict[str, Any]) -> str | None:
    subject = normalize_text(parsed.get("subject"))
    if not subject or subject == raw_cell:
        return None
    if _looks_like_non_lesson(subject):
        return None
    if looks_like_class_token(subject):
        return None
    return subject


def _grid_candidate_issue(
    index: int,
    candidate: dict[str, Any],
    code: str,
    message: str,
    *,
    field: str | None = None,
    value: Any | None = None,
) -> dict[str, Any]:
    trace = candidate.get("source_trace") if isinstance(candidate.get("source_trace"), dict) else {}
    return {
        "code": code,
        "message": message,
        "candidate_index": index,
        "field": field,
        "value": value,
        "row": trace.get("row"),
        "column": trace.get("column"),
    }


def _parse_corrections(raw: str | None) -> dict[str, Any] | None:
    if not raw:
        return None
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Corrections invalides: JSON attendu.") from exc
    if not isinstance(parsed, dict):
        raise HTTPException(status_code=400, detail="Corrections invalides: objet JSON attendu.")
    return parsed


def _analyze_import_primary(content: bytes, filename: str | None) -> dict[str, Any]:
    try:
        from backend.services.imports.intelligence_adapter import analyze_with_intelligence_brains

        result = analyze_with_intelligence_brains(content, filename=filename)
        result["engine"] = "imports_intelligence_brains"
        return result
    except Exception as exc:
        result = _analyze_content(content, filename)
        result["engine"] = "fallback"
        diagnostics = result.setdefault("diagnostics", [])
        if isinstance(diagnostics, list):
            diagnostics.append(
                {
                    "severity": "warning",
                    "code": "intelligence_engine_failed",
                    "message": str(exc),
                }
            )
            result["needs_human_review"] = True
        return result


def _analyze_content(content: bytes, filename: str | None) -> dict[str, Any]:
    file_type = _file_type(filename)
    if not content:
        return _blocked_analysis(filename, file_type, "empty_file", "Fichier vide.")
    if file_type == "csv":
        return _analyze_csv(content, filename)
    if file_type == "xlsx":
        table_result = _analyze_xlsx_requirement_tables(content, filename)
        if _has_importable_data(table_result.get("summary", {}), table_result.get("normalized_preview", {})):
            return table_result
        preview = preview_excel_schedule(content, filename=filename)
        if preview.get("errors"):
            result = _analyze_xlsx_table(content, filename)
            if result["status"] != "blocked":
                return result
            return _analysis_from_diagnostic(filename, file_type, preview["errors"])
        return _excel_preview_to_analysis(preview, filename=filename)
    return _blocked_analysis(filename, file_type, "unsupported_for_now", "Format non supporté pour l'analyse d'import.")


def _analyze_xlsx_requirement_tables(content: bytes, filename: str | None) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _blocked_analysis(filename, "xlsx", "xlsx_reader_unavailable", "Le lecteur XLSX n'est pas disponible.")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return _blocked_analysis(filename, "xlsx", "xlsx_read_failed", "Fichier Excel vide, corrompu ou non supporté.")

    all_rows = 0
    sheet_profiles = []
    sheet_classifications = []
    mapped_rows: list[dict[str, Any]] = []
    diagnostics: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            non_empty_rows = [row for row in rows if any(_clean(value) for value in row)]
            all_rows += max(0, len(non_empty_rows) - 1)
            header = _detect_requirement_header(non_empty_rows)
            if header is None:
                sheet_profiles.append({"sheet_name": sheet.title, "parser_used": "mvp_v1_requirements_table", "confidence": 0.0})
                sheet_classifications.append({"sheet_name": sheet.title, "sheet_type": "unknown_review", "confidence": 0.0, "needs_human_review": True})
                continue
            header_index, header_row, column_roles, confidence = header
            sheet_profiles.append({"sheet_name": sheet.title, "parser_used": "mvp_v1_requirements_table", "confidence": confidence})
            sheet_classifications.append({"sheet_name": sheet.title, "sheet_type": "requirements_table", "confidence": confidence, "needs_human_review": confidence < 0.75})
            for row_offset, row in enumerate(non_empty_rows[header_index + 1 :], start=header_index + 2):
                mapped: dict[str, Any] = {"source_sheet": sheet.title, "source_row": row_offset}
                for index, value in enumerate(row):
                    role = column_roles.get(index)
                    if not role:
                        continue
                    mapped[role] = _parse_cell_for_role(role, value)
                for key in ("class_name", "teacher_name", "subject_name", "weekly_hours"):
                    mapped.setdefault(key, None)
                if not any(mapped.get(key) for key in ("class_name", "teacher_name", "subject_name", "weekly_hours")):
                    continue
                mapped_rows.append(mapped)
    finally:
        workbook.close()

    classes = _unique(item.get("class_name") for item in mapped_rows)
    teachers = _unique(item.get("teacher_name") for item in mapped_rows)
    subjects = _unique(_canonical_subject(item.get("subject_name")) for item in mapped_rows)
    requirements = []
    for item in mapped_rows:
        requirement = {**item, "subject_name": _canonical_subject(item.get("subject_name"))}
        requirements.append(requirement)
        if not requirement.get("class_name"):
            diagnostics.append(_diagnostic("blocking", "missing_class", "Certaines lignes n'ont pas de classe détectée."))
        if not requirement.get("subject_name"):
            diagnostics.append(_diagnostic("warning", "missing_subject", "Certaines lignes n'ont pas de matière détectée."))
        if not requirement.get("teacher_name"):
            diagnostics.append(_diagnostic("warning", "missing_teacher", "Certaines lignes n'ont pas de professeur détecté."))
        if requirement.get("weekly_hours") is None:
            diagnostics.append(_diagnostic("warning", "missing_or_invalid_hours", "Certaines lignes n'ont pas de volume horaire lisible."))

    diagnostics = _dedupe_diagnostics(diagnostics)
    status = "ok" if requirements and not any(item.get("severity") == "blocking" for item in diagnostics) else "needs_review" if requirements else "blocked"
    result = _analysis_payload(
        filename=filename,
        file_type="xlsx",
        status=status,
        confidence=0.82 if status == "ok" else 0.58 if requirements else 0.0,
        summary={
            "sheets_count": len(sheet_profiles),
            "rows_count": all_rows,
            "importable_rows": len(requirements),
            "requirements_count": len(requirements),
            "classes_count": len(classes),
            "teachers_count": len(teachers),
            "subjects_count": len(subjects),
        },
        normalized_preview={
            "classes": [{"name": name} for name in classes],
            "teachers": [{"name": name} for name in teachers],
            "subjects": [{"name": name} for name in subjects],
            "requirements": requirements,
            "constraints": [],
            "availability": [],
            "source_trace": [],
        },
        diagnostics=diagnostics,
    )
    result["engine_used"] = "mvp_v1_requirements_table"
    result["sheet_profiles"] = sheet_profiles
    result["sheet_classifications"] = sheet_classifications
    return result


def _analyze_csv(content: bytes, filename: str | None) -> dict[str, Any]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.DictReader(StringIO(text), dialect=dialect))
    if not rows or not rows[0]:
        return _blocked_analysis(filename, "csv", "empty_csv", "CSV vide ou sans en-têtes lisibles.")

    mapped_rows = [_map_csv_row(row) for row in rows]
    classes = _unique(item.get("class_name") for item in mapped_rows)
    teachers = _unique(item.get("teacher_name") for item in mapped_rows)
    subjects = _unique(item.get("subject_name") for item in mapped_rows)
    requirements = [item for item in mapped_rows if item.get("class_name") or item.get("subject_name") or item.get("teacher_name")]
    diagnostics = []
    if not requirements:
        diagnostics.append(_diagnostic("warning", "no_requirements", "Aucune ligne de besoin exploitable détectée."))
    if any(not item.get("teacher_name") for item in requirements):
        diagnostics.append(_diagnostic("warning", "missing_teacher", "Certaines lignes n'ont pas de professeur détecté."))
    return _analysis_payload(
        filename=filename,
        file_type="csv",
        status="ok" if requirements else "needs_review",
        confidence=0.75 if requirements else 0.35,
        summary={
            "rows_count": len(rows),
            "requirements_count": len(requirements),
            "classes_count": len(classes),
            "teachers_count": len(teachers),
            "subjects_count": len(subjects),
        },
        normalized_preview={
            "classes": [{"name": name} for name in classes],
            "teachers": [{"name": name} for name in teachers],
            "subjects": [{"name": name} for name in subjects],
            "requirements": requirements,
            "constraints": [],
            "availability": [],
            "source_trace": [],
        },
        diagnostics=diagnostics,
    )


def _analyze_xlsx_table(content: bytes, filename: str | None) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _blocked_analysis(filename, "xlsx", "xlsx_reader_unavailable", "Le lecteur XLSX n'est pas disponible.")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return _blocked_analysis(filename, "xlsx", "xlsx_read_failed", "Fichier Excel vide, corrompu ou non supporté.")

    try:
        best: tuple[int, str, list[Any], list[list[Any]]] | None = None
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            for index, row in enumerate(rows[:20]):
                roles = [_header_role(str(value)) for value in row]
                score = sum(1 for role in roles if role)
                if score >= 2 and (best is None or score > sum(1 for role in [_header_role(str(value)) for value in best[2]] if role)):
                    best = (index, sheet.title, row, rows)
        if best is None:
            return _blocked_analysis(filename, "xlsx", "no_table_detected", "Aucun tableau de besoins exploitable détecté.")

        header_index, sheet_name, header_row, rows = best
        column_roles = {index: _header_role(str(value)) for index, value in enumerate(header_row)}
        mapped_rows = []
        for row in rows[header_index + 1 :]:
            mapped = {}
            for index, value in enumerate(row):
                role = column_roles.get(index)
                if role:
                    mapped[role] = _clean(value)
            if any(mapped.values()):
                mapped_rows.append(mapped)
    finally:
        workbook.close()

    classes = _unique(item.get("class_name") for item in mapped_rows)
    teachers = _unique(item.get("teacher_name") for item in mapped_rows)
    subjects = _unique(item.get("subject_name") for item in mapped_rows)
    requirements = [item for item in mapped_rows if item.get("class_name") or item.get("subject_name") or item.get("teacher_name")]
    diagnostics = []
    if not requirements:
        diagnostics.append(_diagnostic("warning", "no_requirements", "Aucune ligne de besoin exploitable détectée."))
    result = _analysis_payload(
        filename=filename,
        file_type="xlsx",
        status="ok" if requirements else "needs_review",
        confidence=0.72 if requirements else 0.35,
        summary={
            "sheets_count": len(rows) and 1,
            "requirements_count": len(requirements),
            "classes_count": len(classes),
            "teachers_count": len(teachers),
            "subjects_count": len(subjects),
        },
        normalized_preview={
            "classes": [{"name": name} for name in classes],
            "teachers": [{"name": name} for name in teachers],
            "subjects": [{"name": name} for name in subjects],
            "requirements": requirements,
            "constraints": [],
            "availability": [],
            "source_trace": [],
        },
        diagnostics=diagnostics,
    )
    result["sheet_profiles"] = [{"sheet_name": sheet_name, "parser_used": "tracked_xlsx_table"}]
    result["sheet_classifications"] = [{"sheet_name": sheet_name, "sheet_type": "requirements_table"}]
    return result


def _excel_preview_to_analysis(preview: dict[str, Any], filename: str | None, legacy: bool = False) -> dict[str, Any]:
    counts = preview.get("counts", {})
    diagnostics = [_diagnostic("warning", "excel_warning", item) for item in preview.get("warnings", [])]
    payload = _analysis_payload(
        filename=filename,
        file_type="xlsx",
        status="ok",
        confidence=0.7 if preview.get("can_commit") else 0.45,
        summary={
            "sheets_count": 1 if preview.get("sheet_name") else 0,
            "requirements_count": counts.get("lessons", 0),
            "classes_count": counts.get("classes", 0),
            "teachers_count": counts.get("teachers", 0),
            "subjects_count": counts.get("subjects", 0),
        },
        normalized_preview={
            "classes": [{"name": name} for name in preview.get("classes", [])],
            "teachers": [{"name": name} for name in preview.get("teachers", [])],
            "subjects": [{"name": name} for name in preview.get("subjects", [])],
            "requirements": preview.get("lessons", []),
            "constraints": [],
            "availability": [],
            "source_trace": [],
        },
        diagnostics=diagnostics,
        import_id=preview.get("import_id"),
    )
    payload["sheet_profiles"] = [{"sheet_name": preview.get("sheet_name"), "parser_used": preview.get("parser_used")}]
    payload["sheet_classifications"] = [{"sheet_name": preview.get("sheet_name"), "sheet_type": "schedule_grid"}]
    payload["can_commit"] = bool(preview.get("can_commit")) and payload["can_commit"]
    payload["can_apply"] = bool(preview.get("can_commit")) and payload["can_apply"]
    if legacy:
        payload["engine_used"] = "v1"
    return payload


def _analysis_from_diagnostic(filename: str | None, file_type: str, messages: list[str]) -> dict[str, Any]:
    return _analysis_payload(
        filename=filename,
        file_type=file_type,
        status="blocked",
        confidence=0.0,
        summary={"requirements_count": 0},
        normalized_preview=_empty_preview(),
        diagnostics=[_diagnostic("blocking", "unsupported_or_unreadable", message) for message in messages],
    )


def _blocked_analysis(filename: str | None, file_type: str, code: str, message: str) -> dict[str, Any]:
    return _analysis_payload(
        filename=filename,
        file_type=file_type,
        status="blocked",
        confidence=0.0,
        summary={"requirements_count": 0},
        normalized_preview=_empty_preview(),
        diagnostics=[_diagnostic("blocking", code, message)],
    )


def _analysis_payload(
    *,
    filename: str | None,
    file_type: str,
    status: str,
    confidence: float,
    summary: dict[str, Any],
    normalized_preview: dict[str, Any],
    diagnostics: list[dict[str, Any]],
    import_id: str | None = None,
) -> dict[str, Any]:
    diagnostics = _with_empty_data_diagnostic(summary, normalized_preview, diagnostics)
    has_data = _has_importable_data(summary, normalized_preview)
    is_blocked = status == "blocked" or not has_data or any(item.get("severity") == "blocking" for item in diagnostics)
    can_commit = status == "ok" and not is_blocked
    result = {
        "import_id": import_id or _import_id(filename, file_type, summary, normalized_preview),
        "filename": filename,
        "status": "blocked" if is_blocked else status,
        "file_type": file_type,
        "confidence": 0.0 if is_blocked and not has_data else confidence,
        "confidence_score": 0.0 if is_blocked and not has_data else confidence,
        "summary": summary,
        "sheet_profiles": [],
        "sheet_classifications": [],
        "normalized_preview": normalized_preview,
        "diagnostics": diagnostics,
        "human_review": _human_review(diagnostics),
        "needs_human_review": bool(diagnostics) or is_blocked,
        "can_apply": can_commit,
        "can_commit": can_commit,
    }
    return result


def _with_empty_data_diagnostic(
    summary: dict[str, Any],
    preview: dict[str, Any],
    diagnostics: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if _has_importable_data(summary, preview):
        return diagnostics
    if any(item.get("code") == "no_importable_data" for item in diagnostics):
        return diagnostics
    return [
        *diagnostics,
        _diagnostic("blocking", "no_importable_data", "Aucune donnée importable n'a été détectée."),
    ]


def _has_importable_data(summary: dict[str, Any], preview: dict[str, Any]) -> bool:
    requirements_count = int(summary.get("requirements_count") or summary.get("importable_rows") or 0)
    classes_count = int(summary.get("classes_count") or summary.get("detected_classes") or 0)
    teachers_count = int(summary.get("teachers_count") or summary.get("detected_teachers") or 0)
    subjects_count = int(summary.get("subjects_count") or summary.get("detected_subjects") or 0)
    if requirements_count > 0 or any(count > 0 for count in (classes_count, teachers_count, subjects_count)):
        return True
    return any(as_list for as_list in (
        preview.get("requirements") or [],
        preview.get("classes") or [],
        preview.get("teachers") or [],
        preview.get("subjects") or [],
    ))


def _map_csv_row(row: dict[str, Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for key, value in row.items():
        role = _header_role(key)
        if role:
            mapped[role] = _clean(value)
    return mapped


def _detect_requirement_header(rows: list[list[Any]]) -> tuple[int, list[Any], dict[int, str], float] | None:
    best: tuple[int, list[Any], dict[int, str], float] | None = None
    for row_index, row in enumerate(rows[:20]):
        column_roles = {index: _header_role(value) for index, value in enumerate(row)}
        roles = {role for role in column_roles.values() if role}
        business_roles = roles & {"class_name", "subject_name", "teacher_name", "weekly_hours"}
        if len(business_roles) < 2:
            continue
        score = len(business_roles) / 4
        if {"class_name", "subject_name"}.issubset(business_roles):
            score += 0.2
        confidence = min(1.0, round(0.35 + score * 0.65, 2))
        if best is None or confidence > best[3]:
            best = (row_index, row, {index: role for index, role in column_roles.items() if role}, confidence)
    return best


def _header_role(value: str | None) -> str | None:
    normalized = _fold_header(value)
    tokens = {token for token in re_split_header(normalized) if token}
    if normalized in {"classe", "class", "class name", "niveau", "groupe", "כיתה"} or tokens & {"classe", "class", "niveau", "groupe", "כיתה"}:
        return "class_name"
    if normalized in {"matiere", "subject", "subject name", "discipline", "course", "מקצוע"} or tokens & {"matiere", "subject", "discipline", "course", "מקצוע"}:
        return "subject_name"
    if normalized in {"professeur", "prof", "teacher", "teacher name", "enseignant", "מורה", "רב"} or tokens & {"professeur", "prof", "teacher", "enseignant", "מורה", "רב"}:
        return "teacher_name"
    if normalized in {"heures", "hours", "weekly hours", "volume hebdo", "volume horaire", "weekly_hours", "שעות"} or tokens & {"heures", "hours", "שעות"}:
        return "weekly_hours"
    if normalized in {"jour", "day", "יום"} or tokens & {"jour", "day", "יום"}:
        return "day"
    if normalized in {"heure", "time", "creneau", "slot", "שעה"} or tokens & {"heure", "time", "creneau", "slot", "שעה"}:
        return "slot"
    return None


def _parse_cell_for_role(role: str, value: Any) -> Any:
    if role == "weekly_hours":
        return _parse_number(value)
    return _clean(value) or None


def _parse_number(value: Any) -> float | int | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    text = _clean(value).replace(",", ".")
    match = re_search_number(text)
    if not match:
        return None
    number = float(match)
    return int(number) if number.is_integer() else number


def _canonical_subject(value: Any) -> str:
    subject = _clean(value)
    folded = _fold_header(subject)
    aliases = {
        "maths": "Mathématiques",
        "mathematiques": "Mathématiques",
        "math": "Mathématiques",
    }
    return aliases.get(folded, subject)


def _create_missing_entities(preview: dict[str, Any], store: SchedulerRepository) -> dict[str, int]:
    created = {"classes": 0, "teachers": 0, "subjects": 0}
    existing_classes = {_clean(item.name).casefold() for item in store.classes}
    existing_teachers = {_clean(item.name).casefold() for item in store.teachers}
    existing_subjects = {_clean(item.name).casefold() for item in store.subjects}
    for item in preview.get("classes", []):
        name = _clean(item.get("name"))
        if name and name.casefold() not in existing_classes:
            store.add_class(name)
            existing_classes.add(name.casefold())
            created["classes"] += 1
    for item in preview.get("subjects", []):
        name = _clean(item.get("name"))
        if name and name.casefold() not in existing_subjects:
            store.add_subject(name, 1)
            existing_subjects.add(name.casefold())
            created["subjects"] += 1
    for item in preview.get("teachers", []):
        name = _clean(item.get("name"))
        if name and name.casefold() not in existing_teachers:
            store.add_teacher(name, [])
            existing_teachers.add(name.casefold())
            created["teachers"] += 1
    return created


def _file_type(filename: str | None) -> str:
    suffix = (filename or "").rsplit(".", 1)[-1].lower() if "." in (filename or "") else ""
    if suffix in {"csv", "xlsx"}:
        return suffix
    if suffix in {"xlsm", "xls"}:
        return "xlsx"
    return suffix or "unknown"


def _is_excel_filename(filename: str | None) -> bool:
    return (filename or "").lower().endswith((".xlsx", ".xlsm"))


def _empty_preview() -> dict[str, list[Any]]:
    return {"classes": [], "teachers": [], "subjects": [], "requirements": [], "constraints": [], "availability": [], "source_trace": []}


def _diagnostic(severity: str, code: str, message: str) -> dict[str, Any]:
    return {"severity": severity, "code": code, "message": message}


def _dedupe_diagnostics(diagnostics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[Any, ...]] = set()
    result: list[dict[str, Any]] = []
    for item in diagnostics:
        key = (item.get("severity"), item.get("code"), item.get("message"))
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _human_review(diagnostics: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [{"question": item["message"], "code": item["code"]} for item in diagnostics if item.get("severity") in {"blocking", "warning"}]


def _import_id(filename: str | None, file_type: str, summary: dict[str, Any], preview: dict[str, Any]) -> str:
    digest = hashlib.sha1(repr((filename, file_type, summary, preview)).encode("utf-8")).hexdigest()[:12]
    return f"analysis_{digest}"


def _unique(values: Any) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = _clean(value)
        key = cleaned.casefold()
        if cleaned and key not in seen:
            seen.add(key)
            result.append(cleaned)
    return result


def _clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def _fold_header(value: Any) -> str:
    cleaned = _clean(value).casefold()
    decomposed = unicodedata.normalize("NFKD", cleaned)
    without_accents = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[\s_\-]+", " ", without_accents).strip()


def re_split_header(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[/|:()]+", value) if item.strip()]


def re_search_number(value: str) -> str | None:
    match = re.search(r"-?\d+(?:\.\d+)?", value)
    return match.group(0) if match else None
