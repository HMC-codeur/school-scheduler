from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest

pytest.importorskip("httpx")
from fastapi.testclient import TestClient

from backend.data.store import get_store
from backend.main import app
from backend.services.imports.excel_mvp.draft_store import clear_import_drafts


client = TestClient(app)


def setup_function() -> None:
    get_store().clear_all()
    clear_import_drafts()


def _xlsx(rows: list[list[object]]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _workbook_bytes(sheets: dict[str, list[list[object]]]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    first = True
    for name, rows in sheets.items():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = name
        for row in rows:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _merged_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Planning FR"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Planning semaine"
    sheet.append(["Heure", "Lundi", "Mardi", "Mercredi"])
    sheet.append(["08:00-08:45", "Math 7A\nProfesseur: Cohen\nSalle: 101", "", "Français 8B\nProfesseur: Levi"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _analyze(rows: list[list[object]]) -> dict:
    response = client.post(
        "/imports/excel/analyze",
        files={"file": ("requirements.xlsx", _xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    return response.json()


def _analyze_bytes(content: bytes, filename: str = "requirements.xlsx") -> dict:
    response = client.post(
        "/imports/excel/analyze",
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    return response.json()


def _analyze_bytes_with_corrections(content: bytes, corrections: dict, filename: str = "requirements.xlsx") -> dict:
    response = client.post(
        "/imports/excel/analyze",
        data={"corrections": json.dumps(corrections)},
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    return response.json()


def test_excel_mvp_analyze_without_file_returns_clear_error() -> None:
    response = client.post("/imports/excel/analyze")

    assert response.status_code == 400
    assert response.json()["detail"] == "Aucun fichier Excel reçu."


def test_excel_mvp_analyze_rejects_non_excel_extension_with_clear_error() -> None:
    response = client.post(
        "/imports/excel/analyze",
        files={"file": ("requirements.txt", b"not excel", "text/plain")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Format invalide. Utilisez un fichier .xlsx ou .xlsm."


def test_excel_mvp_analyze_rejects_unreadable_excel_with_clear_error() -> None:
    response = client.post(
        "/imports/excel/analyze",
        files={"file": ("requirements.xlsx", b"not excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Impossible de lire ce fichier Excel. Essayez de le réenregistrer en .xlsx depuis Excel ou Google Sheets."
    assert "expected <class" not in response.text


def _requirements(count: int, header: list[object] | None = None) -> list[list[object]]:
    rows = [header or ["Classe", "Matière", "Professeur", "Heures"]]
    rows.extend([[f"{7 + index % 3}A", f"Subject {index}", f"Teacher {index % 5}", 2] for index in range(1, count + 1)])
    return rows


def test_excel_mvp_reads_all_10_data_rows() -> None:
    payload = _analyze(_requirements(10))

    assert payload["reader_used"]
    assert payload["reader_attempts"]
    assert payload["summary"]["data_rows_detected"] == 10
    assert payload["summary"]["requirements_detected"] == 10
    assert len(payload["extracted_entities"]["requirements"]) == 10


def test_excel_analyze_reader_crash_is_sanitized_and_fallback_can_succeed(monkeypatch) -> None:
    import backend.services.imports.excel_mvp.pipeline as pipeline

    def fake_read_excel_with_fallback(content: bytes, filename: str | None = None) -> dict:
        return {
            "reader_used": "fallback_reader",
            "reader_warnings": [],
            "reader_attempts": [
                {
                    "reader_name": "openpyxl_normal",
                    "success": False,
                    "error": "openpyxl n'a pas pu lire les styles du fichier.",
                    "quality_score": 0,
                    "warnings": [],
                    "sheets_count": 0,
                },
                {
                    "reader_name": "fallback_reader",
                    "success": True,
                    "error": None,
                    "quality_score": 90,
                    "warnings": [],
                    "sheets_count": 1,
                },
            ],
            "sheets": [
                {
                    "sheet_name": "דרישות",
                    "rows": [
                        {"row_index": 1, "values": ["Classe", "Matière", "Professeur", "Heures"]},
                        {"row_index": 2, "values": ["7A", "Math", "Cohen", "4"]},
                    ],
                    "max_row": 2,
                    "max_column": 4,
                    "merged_ranges": [],
                    "warnings": [],
                }
            ],
            "quality_score": 90,
        }

    monkeypatch.setattr(pipeline, "read_excel_with_fallback", fake_read_excel_with_fallback)
    response = client.post(
        "/imports/excel/analyze",
        files={"file": ("style-broken.xlsx", b"not-used", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert "expected <class" not in response.text
    payload = response.json()
    assert payload["reader_used"] == "fallback_reader"
    assert payload["reader_attempts"][0]["error"] == "openpyxl n'a pas pu lire les styles du fichier."
    assert payload["summary"]["requirements_detected"] == 1


def test_excel_analyze_all_reader_failures_never_return_raw_style_error(monkeypatch) -> None:
    import backend.services.imports.excel_mvp.pipeline as pipeline
    from backend.services.imports.excel_readers import ExcelReadError

    def fake_read_excel_with_fallback(content: bytes, filename: str | None = None) -> dict:
        raise ExcelReadError(
            attempts=[
                {
                    "reader_name": "openpyxl_normal",
                    "success": False,
                    "error": "openpyxl n'a pas pu lire les styles du fichier.",
                    "quality_score": 0,
                    "warnings": [],
                    "sheets_count": 0,
                }
            ]
        )

    monkeypatch.setattr(pipeline, "read_excel_with_fallback", fake_read_excel_with_fallback)
    response = client.post(
        "/imports/excel/analyze",
        files={"file": ("style-broken.xlsx", b"not-used", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Impossible de lire ce fichier Excel. Essayez de le réenregistrer en .xlsx depuis Excel ou Google Sheets."
    assert "expected <class" not in response.text


def test_excel_mvp_reads_all_30_data_rows() -> None:
    payload = _analyze(_requirements(30))

    assert payload["summary"]["data_rows_detected"] == 30
    assert payload["summary"]["imported_rows_count"] == 30
    assert payload["summary"]["requirements_detected"] == 30
    assert len(payload["extracted_entities"]["requirements"]) == 30
    assert payload["sheets"][0]["detected_format"] == "requirements_table"


def test_excel_mvp_detects_schedule_grid_sheet() -> None:
    payload = _analyze_bytes(
        _workbook_bytes(
            {
                "מערכת שעות": [
                    ["שיעור", "ראשון", "שני", "שלישי"],
                    ["08:00-08:45", "מתמטיקה ז1\nמורה: כהן\nחדר: 101", "", "אנגלית ח2\nמורה: לוי\nחדר: 202"],
                    ["09:00-09:45", "מדעים ז1\nמורה: כהן\nחדר: מעבדה", "תנך ט3\nמורה: ישראלי", ""],
                ]
            }
        ),
        filename="מערכת שעות.xlsx",
    )

    sheet = payload["sheets"][0]
    lessons = sheet["extracted_entities"]["scheduled_lessons"]
    assert sheet["detected_format"] == "schedule_grid"
    assert sheet["confidence"] >= 0.7
    assert sheet["summary"]["scheduled_lessons_detected"] == 4
    assert lessons[0]["day"] == "ראשון"
    assert lessons[0]["start_time"] == "08:00"
    assert lessons[0]["subject"] == "מתמטיקה"
    assert lessons[0]["class_name"] == "ז1"
    assert lessons[0]["teacher_name"] == "כהן"
    assert lessons[0]["room_name"] == "101"
    assert payload["workbook_summary"]["detected_formats"] == ["schedule_grid"]


def test_excel_intelligence_response_observes_workbook_and_patterns() -> None:
    payload = _analyze_bytes(
        _workbook_bytes(
            {
                "מערכת שעות": [
                    ["שיעור", "ראשון", "שני", "שלישי"],
                    ["08:00-08:45", "גמרא עיון יב1\nמורה:קרוסקל שאול\nחדר:כיתת אם י\"ב", "", ""],
                ]
            }
        ),
        filename="real-school.xlsx",
    )

    assert payload["needs_human_validation"] is True
    assert payload["validation_questions"]
    assert payload["normalized_entities"]["scheduled_lessons"][0]["teacher_name"] == "קרוסקל שאול"
    observed = payload["workbook_summary"]["observed"]
    assert observed["sheets_count"] == 1
    assert "he" in observed["detected_languages"]
    sheet = payload["sheets"][0]
    pattern_names = {item["name"] for item in sheet["patterns_detected"]}
    possible_types = {item["type"] for item in sheet["possible_types"]}
    assert {"days_in_columns", "time_slots_in_rows", "teacher_prefix", "room_prefix"}.issubset(pattern_names)
    assert "schedule_grid" in possible_types


def test_excel_intelligence_supports_merged_cells_in_observation() -> None:
    payload = _analyze_bytes(_merged_workbook_bytes(), filename="planning.xlsx")

    sheet = payload["sheets"][0]
    assert sheet["observation"]["merged_cells_count"] >= 1
    assert sheet["detected_format"] == "schedule_grid"
    assert "schedule_grid" in {item["type"] for item in sheet["possible_types"]}


def test_excel_intelligence_unknown_workbook_asks_validation() -> None:
    payload = _analyze([["Export libre"], ["note", "foo"], ["bar", "baz"]])

    assert payload["needs_human_validation"] is True
    assert payload["validation_questions"]
    assert payload["normalized_entities"]["requirements"] == []


def test_excel_mvp_analyzes_multiple_sheet_formats() -> None:
    payload = _analyze_bytes(
        _workbook_bytes(
            {
                "דרישות": [["Classe", "Matière", "Professeur", "Heures"], ["7A", "Math", "Cohen", 4]],
                "מערכת שעות": [
                    ["שיעור", "ראשון", "שני"],
                    ["08:00", "Math 7A\nמורה: Cohen\nחדר: 1", ""],
                ],
            }
        )
    )

    formats = {sheet["detected_format"] for sheet in payload["sheets"]}
    assert formats == {"requirements_table", "schedule_grid"}
    assert payload["summary"]["requirements_detected"] == 1
    assert payload["summary"]["scheduled_lessons_detected"] == 1


def test_excel_mvp_summary_uses_full_dataset_not_preview_rows() -> None:
    header = ["Classe", "Matière", "Professeur", "Heures"]
    data_rows = [[f"{7 + index % 3}A", f"Subject {index}", f"Teacher {index % 5}", 2] for index in range(1, 31)]
    rows = [["Export école"], [], header, *data_rows[:15], [], *data_rows[15:]]
    payload = _analyze(rows)

    assert payload["summary"]["total_rows_read"] == 34
    assert payload["summary"]["detected_header_row"] == 3
    assert payload["summary"]["data_rows_detected"] == 30
    assert payload["summary"]["imported_rows_count"] == 30
    assert payload["summary"]["ignored_empty_rows"] == 2
    assert len(payload["extracted_entities"]["requirements"]) == 30


def test_excel_mvp_detects_header_after_two_empty_rows() -> None:
    payload = _analyze([[], [], *_requirements(15)])

    assert payload["summary"]["detected_header_row"] == 3
    assert payload["summary"]["data_rows_detected"] == 15


def test_excel_mvp_detects_header_after_title_and_blank() -> None:
    payload = _analyze([["Export école"], [], ["Classe", "Matière", "Professeur", "Heures"], ["7A", "Math", "David", 4]])

    assert payload["summary"]["detected_header_row"] == 3
    assert payload["summary"]["data_rows_detected"] == 1


def test_excel_mvp_does_not_stop_on_empty_rows_in_the_middle() -> None:
    rows = [["Classe", "Matière", "Professeur", "Heures"], ["7A", "Math", "David", 4], [], [], ["8B", "Science", "Miriam", 3]]
    payload = _analyze(rows)

    assert payload["summary"]["data_rows_detected"] == 2
    assert payload["summary"]["ignored_empty_rows"] == 2


def test_excel_mvp_supports_french_columns() -> None:
    payload = _analyze([["Classe", "Matière", "Professeur", "Heures"], ["7A", "Math", "David", 4]])

    fields = {column["mapped_field"] for column in payload["detected_columns"]}
    assert {"class_name", "subject_name", "teacher_name", "weekly_hours"}.issubset(fields)
    assert payload["summary"]["classes_detected"] == 1


def test_excel_mvp_supports_hebrew_columns() -> None:
    payload = _analyze([["כיתה", "מקצוע", "מורה", "שעות"], ["ז1", "חשבון", "הרב כהן", 5]])

    fields = {column["mapped_field"] for column in payload["detected_columns"]}
    assert {"class_name", "subject_name", "teacher_name", "weekly_hours"}.issubset(fields)
    assert payload["summary"]["subjects_detected"] == 1


def test_excel_intelligence_v2_detects_hebrew_requirements_table(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze([["כיתה", "מקצוע", "מורה", "שעות"], ["ז1", "חשבון", "הרב כהן", 5]])

    sheet = payload["sheets"][0]
    assert payload["excel_intelligence_mode"] == "v2"
    assert payload["engine_used"] == "v2"
    assert sheet["detected_format"] == "requirements_table"
    assert sheet["confidence"] >= 90
    assert payload["summary"]["requirements_detected"] == 1
    assert sheet["parser_selection"]["parser_name"] == "requirements_parser"


def test_excel_intelligence_v2_detects_english_requirements_table(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze([["class", "subject", "teacher", "hours"], ["7A", "Math", "Cohen", 4]])

    sheet = payload["sheets"][0]
    assert sheet["detected_format"] == "requirements_table"
    assert sheet["confidence"] >= 90
    assert payload["extracted_entities"]["requirements"][0]["class_name"] == "7A"


def test_excel_intelligence_v2_availability_grid_beats_schedule_grid(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")
    workbook = _workbook_bytes(
        {
            "Teachers availability": [
                ["מורה", "ראשון", "שני", "שלישי"],
                ["כהן", "זמין", "לא זמין", "פנוי"],
                ["לוי", "available", "unavailable", "yes"],
            ]
        }
    )

    payload = _analyze_bytes(workbook, filename="demo_teacher_availability.xlsx")

    sheet = payload["sheets"][0]
    assert sheet["detected_format"] == "availability_grid"
    assert sheet["confidence"] >= 80
    assert payload["extracted_entities"]["teacher_availability"]
    assert payload["extracted_entities"]["scheduled_lessons"] == []
    assert sheet["summary"]["scheduled_lessons_detected"] == 0
    assert sheet["hypotheses"][0]["format"] == "availability_grid"


def test_excel_intelligence_v2_keeps_real_schedule_grid(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")
    workbook = _workbook_bytes(
        {
            "מערכת שעות": [
                ["שיעור", "ראשון", "שני"],
                ["08:00-08:45", "מתמטיקה ז1\nמורה: כהן\nחדר: 101", "אנגלית ח2\nמורה: לוי\nחדר: 202"],
                ["09:00-09:45", "מדעים ז1\nמורה: כהן\nחדר: מעבדה", ""],
            ]
        }
    )

    payload = _analyze_bytes(workbook, filename="demo_schedule_grid.xlsx")

    sheet = payload["sheets"][0]
    assert sheet["detected_format"] == "schedule_grid"
    assert sheet["extracted_entities"]["scheduled_lessons"]
    assert payload["summary"]["scheduled_lessons_detected"] == 3


def test_excel_intelligence_v2_ignores_noisy_sheet(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze([["WhatsApp"], ["Rav Cohen says maybe move Tuesday, ask parents"], ["Reminder: bring forms, no final planning here"]])

    sheet = payload["sheets"][0]
    assert sheet["detected_format"] in {"noisy", "unknown"}
    assert sheet["confidence"] < 40
    assert sheet["extracted_entities"] == {}
    assert sheet["ignored_reason"]
    assert any("Classification Excel v2" in item["title"] for item in sheet["diagnostics"])


def test_excel_intelligence_v2_ignores_metadata_oracle_sheets(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes(
        _workbook_bytes(
            {
                "EXPECTED_IMPORT": [["entity_type", "class", "subject"], ["lesson", "7A", "Math"]],
                "TEST_NOTES": [["note"], ["should never import"]],
                "Sources": [["file", "owner"], ["fixture", "qa"]],
            }
        ),
        filename="metadata.xlsx",
    )

    for sheet in payload["sheets"]:
        codes = {item["code"] for item in sheet["diagnostics"]}
        assert sheet["detected_format"] == "metadata_or_oracle"
        assert sheet["import_action"] == "ignored"
        assert sheet["extracted_entities"].get("scheduled_lessons") == []
        assert "no_class_detected" not in codes
        assert "no_subject_detected" not in codes
    assert payload["summary"]["requirements_detected"] == 0
    assert payload["summary"]["scheduled_lessons_detected"] == 0


def test_excel_intelligence_v2_constraints_text_never_creates_lessons(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes(
        _workbook_bytes(
            {
                "Contraintes libres": [
                    ["Cohen ne travaille pas lundi matin"],
                    ["Math pas plus de 2 heures par jour"],
                    ["רב דוד לא זמין ביום ראשון"],
                ]
            }
        ),
        filename="constraints.xlsx",
    )

    sheet = payload["sheets"][0]
    codes = {item["code"] for item in sheet["diagnostics"]}
    assert sheet["detected_format"] == "constraints_text"
    assert sheet["import_action"] == "candidate_review"
    assert sheet["needs_human_review"] is True
    assert sheet["extracted_entities"]["scheduled_lessons"] == []
    assert sheet["extracted_entities"]["constraint_candidates"]
    assert "no_class_detected" not in codes
    assert "no_subject_detected" not in codes
    assert payload["summary"]["scheduled_lessons_detected"] == 0


def test_excel_intelligence_v2_mixed_lists_are_review_only(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes(
        _workbook_bytes(
            {
                "Listes mélangées": [
                    ["Type", "Nom", "Alias", "Remarque"],
                    ["teacher", "Cohen", "", "ne travaille pas lundi matin"],
                    ["class", "7A", "", ""],
                    ["subject", "Math", "", "pas plus de 2 heures par jour"],
                ]
            }
        ),
        filename="mixed.xlsx",
    )

    sheet = payload["sheets"][0]
    codes = {item["code"] for item in sheet["diagnostics"]}
    assert sheet["detected_format"] == "mixed_list"
    assert sheet["import_action"] == "candidate_review"
    assert sheet["extracted_entities"]["scheduled_lessons"] == []
    assert sheet["extracted_entities"]["teacher_candidates"] == ["Cohen"]
    assert sheet["extracted_entities"]["class_candidates"] == ["7A"]
    assert sheet["extracted_entities"]["subject_candidates"] == ["Math"]
    assert "no_class_detected" not in codes
    assert "no_subject_detected" not in codes


def test_pytest_root_config_ignores_temporary_folders() -> None:
    config = (pytest.__file__,)
    text = open("pytest.ini", encoding="utf-8").read()

    assert config
    assert "testpaths = tests" in text
    assert "pytest-temp" in text
    assert ".tmp" in text


def test_excel_human_override_ignored_prevents_extraction(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes_with_corrections(
        _xlsx([["class", "subject", "teacher", "hours"], ["7A", "Math", "Cohen", 4]]),
        {"sheet_overrides": {"Sheet": {"format": "ignored"}}},
    )

    sheet = payload["sheets"][0]
    assert sheet["detected_format"] == "ignored"
    assert sheet["status"] == "ignored"
    assert sheet["extracted_entities"] == {}
    assert payload["summary"]["requirements_detected"] == 0
    assert sheet["user_correction"]["old_format"] == "requirements_table"
    assert sheet["user_correction"]["new_format"] == "ignored"
    assert sheet["diagnostics"][0]["code"] == "human_correction_applied"


def test_excel_human_override_ambiguous_requirements_with_column_roles(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes_with_corrections(
        _xlsx([["A", "B", "C", "D"], ["7A", "Math", "Cohen", 4]]),
        {
            "sheet_overrides": {
                "Sheet": {
                    "format": "requirements_table",
                    "column_roles": {"A": "class", "B": "subject", "C": "teacher", "D": "hours"},
                }
            }
        },
    )

    sheet = payload["sheets"][0]
    requirement = payload["extracted_entities"]["requirements"][0]
    assert sheet["detected_format"] == "requirements_table"
    assert sheet["status"] == "ready"
    assert sheet["needs_review"] is False
    assert sheet["column_role_overrides_applied"] is True
    assert requirement["class_name"] == "7A"
    assert requirement["subject_name"] == "Math"
    assert requirement["teacher_name"] == "Cohen"
    assert requirement["weekly_hours"] == 4
    assert sheet["user_correction"]["old_format"] in {"unknown", "noisy", "requirements_table"}
    assert sheet["user_correction"]["new_format"] == "requirements_table"


def test_excel_human_override_availability_grid_extracts_entries(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes_with_corrections(
        _xlsx([["teacher", "Monday", "Tuesday"], ["Cohen", "available", "unavailable"]]),
        {"sheet_overrides": {"Sheet": {"format": "availability_grid"}}},
        filename="availability.xlsx",
    )

    sheet = payload["sheets"][0]
    assert sheet["detected_format"] == "availability_grid"
    assert sheet["status"] == "ready"
    assert payload["extracted_entities"]["teacher_availability"]
    assert payload["summary"]["teacher_availability_detected"] == 2
    assert sheet["diagnostics"][0]["old_format"]


def test_excel_human_override_diagnostics_show_old_and_new_format(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes_with_corrections(
        _xlsx([["WhatsApp"], ["free text only"]]),
        {"sheet_overrides": {"Sheet": {"format": "schedule_grid"}}},
        filename="notes.xlsx",
    )

    sheet = payload["sheets"][0]
    diagnostic_item = sheet["diagnostics"][0]
    assert sheet["user_correction"]["applied"] is True
    assert diagnostic_item["code"] == "human_correction_applied"
    assert diagnostic_item["old_format"] in {"unknown", "noisy"}
    assert diagnostic_item["new_format"] == "schedule_grid"
    assert sheet["status"] == "ready"


def test_excel_intelligence_v2_stress_multi_sheet_does_not_crash(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")
    workbook = _workbook_bytes(
        {
            "demo_requirements_table": [["class", "subject", "teacher", "hours"], ["7A", "Math", "Cohen", 4]],
            "demo_schedule_grid": [["שיעור", "ראשון"], ["08:00", "Math 7A\nמורה: Cohen\nחדר: 1"]],
            "demo_teacher_availability": [["teacher", "Monday", "Tuesday"], ["Cohen", "available", "unavailable"]],
            "notes": [["WhatsApp"], ["free text and reminders only"]],
        }
    )

    payload = _analyze_bytes(workbook, filename="ultra_complex_school_excel_stress_test.xlsx")

    formats = {sheet["sheet_name"]: sheet["detected_format"] for sheet in payload["sheets"]}
    assert formats["demo_requirements_table"] == "requirements_table"
    assert formats["demo_schedule_grid"] == "schedule_grid"
    assert formats["demo_teacher_availability"] == "availability_grid"
    assert formats["notes"] in {"noisy", "unknown", "metadata_or_oracle"}
    assert payload["extracted_entities"]["requirements"]
    assert payload["extracted_entities"]["teacher_availability"]


def _ultra_complex_school_stress_workbook() -> bytes:
    return _workbook_bytes(
        {
            "01 Besoins MIX": [
                ["Export école - besoins horaires 2026"],
                ["Version envoyée par WhatsApp puis corrigée à la main"],
                [],
                ["כיתה / Classe", "מקצוע / Matière", "מורה / Professeur", "שעות / Heures", "Salle préférée", "Notes", "Priorité"],
                ["ז1", "מתמטיקה", "כהן", 5, "101", "garder matin", "haute"],
                ["ח2", "אנגלית", "לוי", 4, "202", "", "moyenne"],
            ],
            "02 מערכת שעות גריד": [
                ["שיעור", "ראשון", "שני", "שלישי"],
                ["08:00-08:45", "מתמטיקה ז1\nמורה: כהן\nחדר: 101", "אנגלית ח2\nמורה: לוי\nחדר: 202", ""],
                ["09:00-09:45", "מדעים ז1\nמורה: כהן\nחדר: מעבדה", "", "עברית ח2\nמורה: לוי\nחדר: 204"],
            ],
            "03 זמינות מורים": [
                ["מורה", "ראשון", "שני", "שלישי", "רביעי"],
                ["כהן", "זמין", "לא זמין", "available", "unavailable"],
                ["לוי", "כן", "לא", "פנוי", "לא פנוי"],
                ["ישראלי", "yes", "no", "זמין", "לא זמין"],
            ],
            "04 Contraintes": [
                ["Constraint Type", "Target", "Day", "Time / Slot", "Value", "Severity", "Comment"],
                ["teacher_unavailable", "בס נחמן", "שלישי", "09:00-10:35", "true", "blocking", "Conflit possible avec maths"],
                ["class_max_daily_hours", "י״א 1", "", "", "8", "important", ""],
                ["subject_prefer_morning", "Mathématiques 5 unités", "", "morning", "true", "recommendation", ""],
                ["room_required", "מדעי המחשב", "", "", "ח. מחשבים", "blocking", ""],
                ["avoid_long_sequence", "קרוסקל שאול", "חמישי", "", "4", "important", ""],
            ],
            "05 Listes mélangées": [
                ["Liste copiée depuis autre fichier", "Liste copiée depuis autre fichier", "Liste copiée depuis autre fichier", "Liste copiée depuis autre fichier"],
                ["Type", "Nom", "Alias", "Email"],
                ["teacher", "בס נחמן", "Bas Nahman", "bas@example.school"],
                ["teacher", "זאפ אפרים", "Zapp Ephraim", "zapp@example.school"],
                ["class", "י״א 1", "יא1", ""],
                ["class", "י״ב 1", "יב1", ""],
                ["subject", "Mathématiques 5 unités", "Math", ""],
                ["subject", "מדעי המחשב", "Computer Science", ""],
            ],
            "06 Unknown Noise": [
                ["WhatsApp notes"],
                ["Version brouillon, parents pas confirmés, ne pas importer"],
                ["TODO: appeler Cohen puis revoir salles"],
            ],
        }
    )


def test_excel_intelligence_v2_ultra_stress_requirements_header_after_noise(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes(_ultra_complex_school_stress_workbook(), filename="ultra_complex_school_excel_stress_test.xlsx")
    sheet = next(item for item in payload["sheets"] if item["sheet_name"] == "01 Besoins MIX")

    assert sheet["detected_format"] == "requirements_table"
    assert sheet["confidence"] >= 90
    assert sheet["summary"]["requirements_detected"] > 0
    assert sheet["summary"]["detected_header_row"] == 4
    assert sheet["status"] == "ready"
    assert sheet["summary"]["valid_requirement_rows"] > 0
    assert sheet["summary"]["hours_parse_success_ratio"] > 0
    assert sheet["requirements_validation"]["status"] == "ready"
    assert any(item["code"] == "business_header_row_detected" for item in sheet["diagnostics"])


def test_excel_intelligence_v2_ultra_stress_constraints_not_ready_requirements(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes(_ultra_complex_school_stress_workbook(), filename="ultra_complex_school_excel_stress_test.xlsx")
    sheet = next(item for item in payload["sheets"] if item["sheet_name"] == "04 Contraintes")

    assert not (sheet["detected_format"] == "requirements_table" and sheet["status"] == "ready")
    assert sheet["detected_format"] in {"unknown", "noisy", "requirements_table", "constraints_text"}
    assert sheet["status"] in {"ignored", "needs_review"}
    assert sheet["summary"].get("valid_requirement_rows", 0) == 0
    assert sheet["summary"]["requirements_detected"] == 0
    assert sheet["extracted_entities"]["requirements"] == []
    if sheet["detected_format"] == "requirements_table":
        assert sheet["requirements_validation"]["parser_guardrail_reason"]
        assert any(item["code"] == "requirements_table_validation" for item in sheet["diagnostics"])
    else:
        assert sheet["import_action"] == "candidate_review"


def test_excel_intelligence_v2_ultra_stress_mixed_lists_not_ready_requirements(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes(_ultra_complex_school_stress_workbook(), filename="ultra_complex_school_excel_stress_test.xlsx")
    sheet = next(item for item in payload["sheets"] if item["sheet_name"] == "05 Listes mélangées")

    assert not (sheet["detected_format"] == "requirements_table" and sheet["status"] == "ready")
    assert sheet["detected_format"] in {"unknown", "noisy", "requirements_table", "mixed_list", "entity_list"}
    assert sheet["status"] in {"ignored", "needs_review"}
    assert sheet["summary"].get("valid_requirement_rows", 0) == 0
    assert sheet["summary"]["requirements_detected"] == 0
    assert sheet["extracted_entities"]["requirements"] == []
    if sheet["detected_format"] == "requirements_table":
        assert "עמודת השעות אינה ניתנת לזיהוי בצורה אמינה" in sheet["requirements_validation"]["negative_evidence"]
    else:
        assert sheet["import_action"] == "candidate_review"


def test_excel_intelligence_v2_ultra_stress_availability_beats_schedule(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes(_ultra_complex_school_stress_workbook(), filename="ultra_complex_school_excel_stress_test.xlsx")
    sheet = next(item for item in payload["sheets"] if item["sheet_name"] == "03 זמינות מורים")
    marker_subjects = {"זמין", "לא זמין", "פנוי", "לא פנוי", "available", "unavailable", "yes", "no", "כן", "לא"}

    assert sheet["detected_format"] == "availability_grid"
    assert sheet["summary"]["availability_entries_detected"] > 0
    assert sheet["summary"]["scheduled_lessons_detected"] == 0
    assert sheet["extracted_entities"]["scheduled_lessons"] == []
    assert payload["extracted_entities"]["teacher_availability"]
    assert not any((lesson.get("subject") or lesson.get("subject_name")) in marker_subjects for lesson in payload["extracted_entities"]["scheduled_lessons"])
    assert any(item["code"] == "availability_markers_detected" for item in sheet["diagnostics"])


def test_excel_analyze_endpoint_default_v1_remains_primary() -> None:
    response = client.post(
        "/imports/excel/analyze",
        files={
            "file": (
                "default-v1.xlsx",
                _ultra_complex_school_stress_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    sheets = {item["sheet_name"]: item for item in payload["sheets"]}
    assert payload["engine_used"] == "v1"
    assert payload.get("excel_intelligence_mode") is None
    assert "03 זמינות מורים" in sheets
    assert "parser_selection" not in sheets["03 זמינות מורים"]


def test_excel_analyze_endpoint_env_v2_uses_real_endpoint_path(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    response = client.post(
        "/imports/excel/analyze",
        files={
            "file": (
                "env-v2.xlsx",
                _ultra_complex_school_stress_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    sheets = {item["sheet_name"]: item for item in payload["sheets"]}
    availability = sheets["03 זמינות מורים"]
    marker_subjects = {"זמין", "לא זמין", "available", "unavailable", "yes", "no", "כן", "לא"}

    assert payload["engine_used"] == "v2"
    assert sheets["01 Besoins MIX"]["detected_format"] == "requirements_table"
    assert sheets["01 Besoins MIX"]["confidence"] >= 90
    assert availability["detected_format"] == "availability_grid"
    assert availability["summary"]["availability_entries_detected"] > 0
    assert availability["summary"]["scheduled_lessons_detected"] == 0
    assert payload["extracted_entities"]["teacher_availability"]
    assert not any((lesson.get("subject") or lesson.get("subject_name")) in marker_subjects for lesson in payload["extracted_entities"]["scheduled_lessons"])


def test_excel_analyze_real_uploaded_ultra_file_v2(monkeypatch) -> None:
    workbook_path = Path(".tmp/ultra_complex_school_excel_stress_test.xlsx")
    if not workbook_path.exists():
        pytest.skip("real uploaded stress workbook is not present")
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    response = client.post(
        "/imports/excel/analyze",
        files={
            "file": (
                workbook_path.name,
                workbook_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    sheets = {item["sheet_name"]: item for item in payload["sheets"]}
    marker_subjects = {"זמין", "לא זמין", "available", "unavailable", "yes", "no", "כן", "לא"}

    assert payload["engine_used"] == "v2"
    assert sheets["01 Besoins MIX"]["detected_format"] == "requirements_table"
    assert sheets["01 Besoins MIX"]["confidence"] >= 90
    assert sheets["03 זמינות מורים"]["detected_format"] == "availability_grid"
    assert len(sheets["03 זמינות מורים"]["extracted_entities"]["teacher_availability"]) > 0
    assert sheets["03 זמינות מורים"]["summary"]["scheduled_lessons_detected"] == 0
    if "04 Contraintes" in sheets:
        assert not (sheets["04 Contraintes"]["detected_format"] == "requirements_table" and sheets["04 Contraintes"]["status"] == "ready")
        assert sheets["04 Contraintes"]["summary"].get("requirements_detected", 0) == 0
    if "05 Listes mélangées" in sheets:
        assert not (sheets["05 Listes mélangées"]["detected_format"] == "requirements_table" and sheets["05 Listes mélangées"]["status"] == "ready")
        assert sheets["05 Listes mélangées"]["summary"].get("requirements_detected", 0) == 0
    assert not any((lesson.get("subject") or lesson.get("subject_name")) in marker_subjects for lesson in payload["extracted_entities"]["scheduled_lessons"])


def test_excel_intelligence_v2_ultra_stress_keeps_real_schedule_and_noise(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    payload = _analyze_bytes(_ultra_complex_school_stress_workbook(), filename="ultra_complex_school_excel_stress_test.xlsx")
    sheets = {item["sheet_name"]: item for item in payload["sheets"]}

    schedule = sheets["02 מערכת שעות גריד"]
    assert schedule["detected_format"] == "schedule_grid"
    assert schedule["summary"]["scheduled_lessons_detected"] >= 3

    constraints = sheets["04 Contraintes"]
    assert constraints["summary"]["requirements_detected"] == 0
    assert constraints["status"] in {"ignored", "needs_review"}

    mixed_lists = sheets["05 Listes mélangées"]
    assert mixed_lists["summary"]["requirements_detected"] == 0
    assert mixed_lists["status"] in {"ignored", "needs_review"}

    noise = sheets["06 Unknown Noise"]
    assert noise["detected_format"] in {"unknown", "noisy"}
    assert noise["confidence"] < 40
    assert noise["extracted_entities"] == {}


def test_excel_intelligence_debug_compare_keeps_v1_as_primary() -> None:
    response = client.post(
        "/imports/excel/analyze?debug_compare=true",
        files={
            "file": (
                "compare.xlsx",
                _workbook_bytes(
                    {
                        "requirements": [["class", "subject", "teacher", "hours"], ["7A", "Math", "Cohen", 4]],
                        "availability": [["teacher", "Monday", "Tuesday"], ["Cohen", "available", "unavailable"]],
                    }
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    compare = payload["excel_intelligence_compare"]
    assert payload.get("excel_intelligence_mode") is None
    assert payload["engine_used"] == "v1"
    assert payload["debug_compare"] is True
    assert payload["primary_result_engine"] == "v1"
    assert compare["primary_mode"] == "v1"
    assert compare["primary_result_engine"] == "v1"
    assert compare["v2_default_enabled"] is False
    assert payload["v1_result"]["engine_used"] == "v1"
    assert payload["v2_result"]["engine_used"] == "v2"
    assert compare["v1_result"]["engine_used"] == "v1"
    assert compare["v2_result"]["engine_used"] == "v2"
    assert "availability_grid" in compare["summary"]["v2_detected_formats"]
    assert compare["summary"]["v2_availability_entries"] >= 1
    assert all("v1_result" in sheet and "v2_result" in sheet for sheet in compare["sheets"])


def test_excel_intelligence_debug_compare_uses_env_v2_as_primary(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_INTELLIGENCE_MODE", "v2")

    response = client.post(
        "/imports/excel/analyze?debug_compare=true",
        files={
            "file": (
                "compare-v2.xlsx",
                _workbook_bytes(
                    {
                        "requirements": [["class", "subject", "teacher", "hours"], ["7A", "Math", "Cohen", 4]],
                        "availability": [["teacher", "Monday", "Tuesday"], ["Cohen", "available", "unavailable"]],
                    }
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    compare = payload["excel_intelligence_compare"]
    assert payload["engine_used"] == "v2"
    assert payload["debug_compare"] is True
    assert payload["primary_result_engine"] == "v2"
    assert compare["primary_mode"] == "v2"
    assert compare["primary_result_engine"] == "v2"
    assert payload["v1_result"]["engine_used"] == "v1"
    assert payload["v2_result"]["engine_used"] == "v2"
    assert any(sheet["detected_format"] == "availability_grid" for sheet in payload["sheets"])


def test_excel_intelligence_debug_compare_reports_v2_noisy() -> None:
    response = client.post(
        "/imports/excel/analyze?debug_compare=true",
        files={
            "file": (
                "notes.xlsx",
                _xlsx([["WhatsApp"], ["Rav Cohen says maybe move Tuesday, ask parents"], ["Reminder: bring forms"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    compare = response.json()["excel_intelligence_compare"]
    sheet = compare["sheets"][0]
    assert sheet["v2_format"] in {"noisy", "unknown"}
    assert sheet["v2_confidence"] < 40
    assert sheet["v2_diagnostic_summary"]


def test_excel_mvp_warns_on_non_numeric_weekly_hours() -> None:
    payload = _analyze([["Classe", "Matière", "Professeur", "Heures"], ["7A", "Math", "David", "beaucoup"]])

    codes = [item["code"] for item in payload["diagnostics"]["warnings"]]
    assert "weekly_hours_not_numeric" in codes
    assert "volume horaire non numérique" in payload["diagnostics"]["warnings"][0]["message"] or any("volume horaire non numérique" in item["message"] for item in payload["diagnostics"]["warnings"])


def test_excel_mvp_blocks_requirement_without_class() -> None:
    payload = _analyze([["Classe", "Matière", "Professeur", "Heures"], ["", "Math", "David", 4]])

    codes = [item["code"] for item in payload["diagnostics"]["blocking"]]
    assert "requirement_missing_class" in codes


def test_excel_mvp_analyze_does_not_mutate_store() -> None:
    store = get_store()
    assert store.classes == []

    _analyze(_requirements(3))

    assert store.classes == []
    assert store.teachers == []
    assert store.subjects == []


def test_excel_mvp_schema_endpoint() -> None:
    response = client.get("/imports/excel/schema")

    assert response.status_code == 200
    assert "class_name" in response.json()["standard_fields"]


def test_import_excel_page_serves_spa_direct_route() -> None:
    response = client.get("/import-excel")

    assert response.status_code == 200
    assert "text/html" in response.headers["content-type"]
    assert response.headers["cache-control"] == "no-store"


def test_excel_mvp_commit_unknown_import_id_returns_404() -> None:
    response = client.post("/imports/excel/missing-import/commit")

    assert response.status_code == 404


def test_excel_mvp_commit_refuses_blocking_diagnostics() -> None:
    payload = _analyze([["Classe", "Matière", "Professeur", "Heures"], ["", "Math", "David", 4]])

    response = client.post(f"/imports/excel/{payload['import_id']}/commit")

    assert response.status_code == 409
    detail = response.json()["detail"]
    assert detail["status"] == "blocked"
    assert detail["blocking_count"] >= 1
    assert get_store().classes == []
    assert get_store().subjects == []
    assert get_store().teachers == []


def test_excel_mvp_commit_valid_excel_adds_classes_teachers_subjects_and_requirements() -> None:
    payload = _analyze(
        [
            ["Classe", "Matière", "Professeur", "Heures"],
            ["7A", "Math", "Cohen", 4],
            ["8B", "Science", "Levi", 3],
        ]
    )

    response = client.post(f"/imports/excel/{payload['import_id']}/commit")

    assert response.status_code == 200
    summary = response.json()["summary"]
    assert summary["classes_added"] == 2
    assert summary["teachers_added"] == 2
    assert summary["subjects_added"] == 2
    assert summary["requirements_added"] == 2
    store = get_store()
    assert {item.name for item in store.classes} == {"7A", "8B"}
    assert {item.name for item in store.teachers} == {"Cohen", "Levi"}
    assert {item.name for item in store.subjects} == {"Math", "Science"}
    assert len(store.import_requirements) == 2


def test_excel_mvp_commit_deduplicates_exact_existing_class() -> None:
    store = get_store()
    store.add_class("7A")
    payload = _analyze([["Classe", "Matière", "Professeur", "Heures"], ["7A", "Math", "Cohen", 4]])

    response = client.post(f"/imports/excel/{payload['import_id']}/commit")

    assert response.status_code == 200
    assert response.json()["summary"]["classes_added"] == 0
    assert response.json()["summary"]["skipped_duplicates"] >= 1
    assert [item.name for item in store.classes] == ["7A"]


def test_excel_mvp_commit_does_not_merge_near_teacher_names() -> None:
    payload = _analyze(
        [
            ["Classe", "Matière", "Professeur", "Heures"],
            ["7A", "Math", "Cohen", 4],
            ["7A", "Math", "Rav Cohen", 4],
        ]
    )

    response = client.post(f"/imports/excel/{payload['import_id']}/commit")

    assert response.status_code == 200
    assert response.json()["summary"]["teachers_added"] == 2
    assert {item.name for item in get_store().teachers} == {"Cohen", "Rav Cohen"}


def test_excel_mvp_commit_modifies_only_after_explicit_call() -> None:
    payload = _analyze(_requirements(2))
    store = get_store()

    assert store.classes == []
    assert store.teachers == []
    assert store.subjects == []
    assert store.import_requirements == []

    response = client.post(f"/imports/excel/{payload['import_id']}/commit")

    assert response.status_code == 200
    assert len(store.classes) == 2
    assert len(store.teachers) == 2
    assert len(store.subjects) == 2
    assert len(store.import_requirements) == 2


def test_legacy_schedule_excel_commit_endpoint_remains_available() -> None:
    response = client.post("/schedule/import/excel/commit", json={"import_id": "missing-import"})

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail["message"] == "Preview expirée ou inconnue. Relancez la preview ou envoyez lessons dans le body."
